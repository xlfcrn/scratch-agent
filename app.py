import os
import re
import base64
import html
import io
import zipfile
import json
import tempfile
from datetime import datetime
from pathlib import Path


import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import time
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font
from PIL import Image as PILImage
from streamlit_paste_button import paste_image_button
from io import BytesIO

APP_VERSION = "v21_teacher_reference_grounded_diagnosis"

# =========================
# 1. 基础配置
# =========================

load_dotenv()


def get_secret_or_env(key: str, default=None):
    """
    优先读取 Streamlit Cloud 的 Secrets；
    如果没有，则读取本地 .env；
    如果都没有，则使用默认值。
    """
    value = None

    try:
        if hasattr(st, "secrets"):
            value = st.secrets.get(key, None)
    except Exception:
        value = None

    if value is None or value == "":
        value = os.environ.get(key, default)

    return value

DEEPSEEK_API_KEY = get_secret_or_env("DEEPSEEK_API_KEY")

if not DEEPSEEK_API_KEY:
    st.error("没有检测到 DEEPSEEK_API_KEY，请先在 .env 文件或 Streamlit Secrets 中配置。")
    st.stop()

client = OpenAI(
    api_key=DEEPSEEK_API_KEY,
    base_url="https://api.deepseek.com"
)

# 图片文字识别 OCR 配置：使用阿里云百炼 DashScope 的 OpenAI 兼容接口
DASHSCOPE_API_KEY = get_secret_or_env("DASHSCOPE_API_KEY")
QWEN_OCR_MODEL = get_secret_or_env("QWEN_OCR_MODEL", "qwen-vl-ocr")
# 作品运行视频分析使用的 Qwen-VL 模型。可在 Secrets 中配置 QWEN_VL_MODEL。
QWEN_VL_MODEL = get_secret_or_env("QWEN_VL_MODEL", "qwen3-vl-flash")
# QWEN_VL_MODEL 同时用于学生截图结构分析与教师端作品运行视频分析。

ocr_client = None
if DASHSCOPE_API_KEY:
    ocr_client = OpenAI(
        api_key=DASHSCOPE_API_KEY,
        base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
    )

st.set_page_config(
    page_title="图形化编程学习助手",
    page_icon="🐱",
    layout="wide",
    initial_sidebar_state="expanded"
)


# =========================
# 1.2 学生端图形化编程知识库
# =========================

APP_DIR = Path(__file__).resolve().parent
KNOWLEDGE_FILE = APP_DIR / "scratch_knowledge.xlsx"
KNOWLEDGE_SHEET_NAME = "知识库"

KNOWLEDGE_REQUIRED_COLUMNS = [
    "知识编号",
    "知识类别",
    "知识点",
    "知识类型",
    "优先级",
    "关键词",
    "典型问法",
    "核心说明",
    "常见问题或现象",
    "可能原因",
    "学生端引导提示",
    "回答边界",
    "检索文本",
    "启用状态",
]


@st.cache_data(show_spinner=False)
def load_knowledge_base(file_path: str, modified_time: float) -> pd.DataFrame:
    """读取固定的学生端知识库。modified_time 用于文件更新后刷新缓存。"""
    _ = modified_time

    df = pd.read_excel(
        file_path,
        sheet_name=KNOWLEDGE_SHEET_NAME,
        dtype=str,
        engine="openpyxl"
    ).fillna("")

    df.columns = [str(column).strip() for column in df.columns]

    missing_columns = [
        column
        for column in KNOWLEDGE_REQUIRED_COLUMNS
        if column not in df.columns
    ]
    if missing_columns:
        raise ValueError("知识库缺少字段：" + "、".join(missing_columns))

    for column in KNOWLEDGE_REQUIRED_COLUMNS:
        df[column] = df[column].astype(str).str.strip()

    enabled_values = {"启用", "是", "1", "true", "True", "TRUE"}
    df = df[df["启用状态"].isin(enabled_values)].copy()

    priority_map = {
        "核心": 3,
        "常用": 2,
        "一般": 1,
    }
    df["_优先级分值"] = (
        df["优先级"]
        .map(priority_map)
        .fillna(1)
        .astype(float)
    )

    return df.reset_index(drop=True)


def normalize_knowledge_text(text: str) -> str:
    """清理检索文本，保留中文、字母和数字。"""
    text = str(text or "").lower()
    return re.sub(
        r"[\s，。！？；：、,.!?;:（）()【】\[\]“”‘’\"'<>《》]+",
        "",
        text
    )


def split_knowledge_keywords(text: str):
    """把关键词单元格拆分为独立关键词。"""
    return [
        item.strip()
        for item in re.split(r"[,，;；、\s]+", str(text or ""))
        if item.strip()
    ]


def make_character_bigrams(text: str):
    """生成连续双字集合，适合当前小规模中文知识库的轻量检索。"""
    normalized = normalize_knowledge_text(text)

    if not normalized:
        return set()

    if len(normalized) == 1:
        return {normalized}

    return {
        normalized[index:index + 2]
        for index in range(len(normalized) - 1)
    }


def calculate_knowledge_similarity(query: str, target: str) -> float:
    """计算两个中文短文本之间的简单相似度。"""
    query_normalized = normalize_knowledge_text(query)
    target_normalized = normalize_knowledge_text(target)

    if not query_normalized or not target_normalized:
        return 0.0

    if min(len(query_normalized), len(target_normalized)) >= 2 and (
        query_normalized in target_normalized
        or target_normalized in query_normalized
    ):
        return 1.0

    query_bigrams = make_character_bigrams(query_normalized)
    target_bigrams = make_character_bigrams(target_normalized)

    if not query_bigrams or not target_bigrams:
        return 0.0

    intersection_count = len(query_bigrams & target_bigrams)
    return (2.0 * intersection_count) / (
        len(query_bigrams) + len(target_bigrams)
    )


def retrieve_knowledge(
    knowledge_df: pd.DataFrame,
    student_question: str,
    current_theme: str = "",
    ocr_text: str = "",
    top_k: int = 3,
) -> pd.DataFrame:
    """
    检索与本轮问题最相关的知识条目。
    当前主题只提供情境，不限制知识库范围。
    """
    if knowledge_df is None or knowledge_df.empty:
        return pd.DataFrame()

    query_parts = [
        str(student_question or "").strip(),
        str(ocr_text or "").strip()[:1800],
    ]
    query_text = "；".join(part for part in query_parts if part)

    if not normalize_knowledge_text(query_text):
        return pd.DataFrame()

    field_weights = {
        "知识点": 5.5,
        "典型问法": 5.0,
        "常见问题或现象": 4.8,
        "可能原因": 3.0,
        "核心说明": 2.3,
        "学生端引导提示": 2.2,
        "检索文本": 1.5,
    }

    normalized_query = normalize_knowledge_text(query_text)
    theme_text = normalize_knowledge_text(current_theme)
    scored_rows = []

    for index, row in knowledge_df.iterrows():
        score = 0.0
        strong_match = False

        for keyword in split_knowledge_keywords(row.get("关键词", "")):
            normalized_keyword = normalize_knowledge_text(keyword)

            if not normalized_keyword:
                continue

            if normalized_keyword in normalized_query:
                score += 8.0
                strong_match = True
            else:
                keyword_similarity = calculate_knowledge_similarity(
                    query_text,
                    keyword
                )
                if keyword_similarity >= 0.62:
                    score += keyword_similarity * 3.0

        for field_name, weight in field_weights.items():
            similarity = calculate_knowledge_similarity(
                query_text,
                str(row.get(field_name, "") or "")
            )
            score += similarity * weight

            if similarity >= 0.5:
                strong_match = True

        # 当前主题只提供很小的情境加成，避免把知识检索限制在某个主题。
        if theme_text and theme_text in normalize_knowledge_text(row.get("检索文本", "")):
            score += 0.3

        # 优先级只参与相关条目之间的排序。
        if strong_match:
            score += float(row.get("_优先级分值", 1)) * 0.3

        if strong_match and score >= 2.5:
            scored_rows.append({
                "index": index,
                "score": score,
            })

    if not scored_rows:
        return pd.DataFrame()

    scored_rows = sorted(
        scored_rows,
        key=lambda item: item["score"],
        reverse=True
    )[:max(1, top_k)]

    selected_indices = [item["index"] for item in scored_rows]
    selected_df = knowledge_df.loc[selected_indices].copy()
    selected_df["_匹配分数"] = [
        round(item["score"], 4)
        for item in scored_rows
    ]

    return selected_df.reset_index(drop=True)


def build_knowledge_context(matched_df: pd.DataFrame) -> str:
    """把检索结果整理为模型能够使用的上下文。"""
    if matched_df is None or matched_df.empty:
        return (
            "本轮没有检索到高度相关的知识条目。"
            "只能依据学生文字和截图中能够确认的信息回答；"
            "如果无法确定，应追问一个关键问题。"
        )

    blocks = []

    for _, row in matched_df.iterrows():
        blocks.append(
            "\n".join([
                f"知识编号：{row.get('知识编号', '')}",
                f"知识类别：{row.get('知识类别', '')}",
                f"知识点：{row.get('知识点', '')}",
                f"知识类型：{row.get('知识类型', '')}",
                f"核心说明：{row.get('核心说明', '')}",
                f"常见问题或现象：{row.get('常见问题或现象', '')}",
                f"可能原因：{row.get('可能原因', '')}",
                f"回答边界：{row.get('回答边界', '')}",
            ])
        )

    return "\n\n---\n\n".join(blocks)


def build_student_knowledge_prompt(
    current_theme: str,
    knowledge_context: str,
) -> str:
    """生成学生端本轮知识库约束。"""
    return f"""
【当前课堂主题】
{current_theme or "未提供"}

【本轮检索到的图形化编程知识】
{knowledge_context}

【知识库使用规则】
1. 知识库用于帮助判断学生当前问题，不要逐字照抄。
2. 当前主题只提供情境，不得把智能体回答范围限制在该主题。
3. 本轮回答的证据优先级为：截图结构分析 > OCR文字 > 知识库 > 一般经验。不得用知识库中的常见原因覆盖截图中已经显示的事实。
4. 在说“缺少某个积木”之前，必须先确认截图结构分析中没有该积木；如果截图中已经存在，不得再建议添加同一个积木。
5. 截图能够明确定位时，只指出一个主要问题，并给出一个关键修改方向。
6. 如果回答里给出坐标、角度、秒数、次数、变量值等具体数值，必须同时用一句话说明这个数值从哪里来，例如来自截图里的当前位置、移动距离、终点位置或两个积木数值的计算。
7. 不得因为知识库列出了多个字段，就把全部原因和建议都告诉学生。
8. 不得直接给出完整作品程序、完整积木组合或完整连接顺序。
9. 知识库没有命中但问题可以确定时，可依据图形化编程基础知识谨慎回答。
10. 知识库没有命中且信息不足时，只追问一个关键问题或请学生补充完整截图。
11. 不得编造知识库内容，也不得声称答案一定正确。
""".strip()


KNOWLEDGE_BASE_DF = pd.DataFrame()
KNOWLEDGE_LOAD_ERROR = ""

if not KNOWLEDGE_FILE.exists():
    KNOWLEDGE_LOAD_ERROR = (
        "未找到 scratch_knowledge.xlsx。"
        "请将它与当前 Python 主程序放在同一目录。"
    )
else:
    try:
        KNOWLEDGE_BASE_DF = load_knowledge_base(
            str(KNOWLEDGE_FILE),
            KNOWLEDGE_FILE.stat().st_mtime,
        )
    except Exception as exc:
        KNOWLEDGE_LOAD_ERROR = f"知识库读取失败：{exc}"



# =========================
# 1.3 教师基础版主题参照
# =========================
# 学生仍然只需要上传截图。教师基础版 .sb3 作为后台参照文件，
# 与当前 Python 主程序放在同一目录即可，不需要学生重复上传。
THEME_REFERENCE_SB3_FILES = {
    "海底世界": "海底世界.sb3",
    "猫捉老鼠": "猫捉老鼠.sb3",
    "牛顿的苹果": "牛顿的苹果.sb3",
    "打地鼠": "打地鼠.sb3",
}


def _scratch_field(block: dict, field_name: str, default="") -> str:
    fields = block.get("fields", {}) if isinstance(block, dict) else {}
    value = fields.get(field_name, default)
    if isinstance(value, list) and value:
        return str(value[0])
    if value is None:
        return str(default)
    return str(value)


def _scratch_primitive(value) -> str:
    """读取 Scratch project.json 中的文字、数字、颜色等原始值。"""
    if isinstance(value, list):
        if len(value) >= 2 and not isinstance(value[1], (list, dict)):
            return str(value[1])
        if len(value) >= 1:
            return str(value[-1])
    if value is None:
        return ""
    return str(value)


def _scratch_input(blocks: dict, block: dict, input_name: str, seen=None) -> str:
    inputs = block.get("inputs", {}) if isinstance(block, dict) else {}
    raw = inputs.get(input_name)
    if not isinstance(raw, list) or len(raw) < 2:
        return ""
    candidate = raw[1]
    if isinstance(candidate, str) and candidate in blocks:
        return _scratch_expression(blocks, candidate, seen=seen)
    if isinstance(candidate, list):
        return _scratch_primitive(candidate)
    if len(raw) >= 3:
        fallback = raw[2]
        if isinstance(fallback, str) and fallback in blocks:
            return _scratch_expression(blocks, fallback, seen=seen)
        if isinstance(fallback, list):
            return _scratch_primitive(fallback)
    return _scratch_primitive(candidate)


def _scratch_expression(blocks: dict, block_id: str, seen=None) -> str:
    seen = set(seen or set())
    if not block_id or block_id in seen:
        return ""
    seen.add(block_id)
    block = blocks.get(block_id, {})
    if not isinstance(block, dict):
        return ""
    opcode = block.get("opcode", "")

    if opcode == "operator_and":
        a = _scratch_input(blocks, block, "OPERAND1", seen)
        b = _scratch_input(blocks, block, "OPERAND2", seen)
        return f"<{a} 与 {b}>"
    if opcode == "operator_or":
        a = _scratch_input(blocks, block, "OPERAND1", seen)
        b = _scratch_input(blocks, block, "OPERAND2", seen)
        return f"<{a} 或 {b}>"
    if opcode == "operator_not":
        a = _scratch_input(blocks, block, "OPERAND", seen)
        return f"不成立<{a}>"
    if opcode in {"operator_equals", "operator_gt", "operator_lt"}:
        a = _scratch_input(blocks, block, "OPERAND1", seen)
        b = _scratch_input(blocks, block, "OPERAND2", seen)
        symbol = {"operator_equals": "=", "operator_gt": ">", "operator_lt": "<"}[opcode]
        return f"<{a} {symbol} {b}>"
    if opcode == "operator_random":
        a = _scratch_input(blocks, block, "FROM", seen)
        b = _scratch_input(blocks, block, "TO", seen)
        return f"在{a}和{b}之间取随机数"
    if opcode == "operator_add":
        return f"({_scratch_input(blocks, block, 'NUM1', seen)} + {_scratch_input(blocks, block, 'NUM2', seen)})"
    if opcode == "operator_subtract":
        return f"({_scratch_input(blocks, block, 'NUM1', seen)} - {_scratch_input(blocks, block, 'NUM2', seen)})"
    if opcode == "operator_multiply":
        return f"({_scratch_input(blocks, block, 'NUM1', seen)} × {_scratch_input(blocks, block, 'NUM2', seen)})"
    if opcode == "operator_divide":
        return f"({_scratch_input(blocks, block, 'NUM1', seen)} ÷ {_scratch_input(blocks, block, 'NUM2', seen)})"
    if opcode == "sensing_touchingobject":
        obj = _scratch_input(blocks, block, "TOUCHINGOBJECTMENU", seen) or _scratch_field(block, "TOUCHINGOBJECTMENU")
        return f"碰到{obj}？"
    if opcode == "sensing_touchingobjectmenu":
        return _scratch_field(block, "TOUCHINGOBJECTMENU")
    if opcode == "sensing_mousedown":
        return "按下鼠标？"
    if opcode == "sensing_keypressed":
        key = _scratch_input(blocks, block, "KEY_OPTION", seen) or _scratch_field(block, "KEY_OPTION")
        return f"按下{key}键？"
    if opcode == "sensing_keyoptions":
        return _scratch_field(block, "KEY_OPTION")
    if opcode == "data_variable":
        return _scratch_field(block, "VARIABLE")
    if opcode == "motion_xposition":
        return "x坐标"
    if opcode == "motion_yposition":
        return "y坐标"
    if opcode == "motion_direction":
        return "方向"
    if opcode == "looks_costumenumbername":
        return "造型编号"
    if opcode == "sensing_mousex":
        return "鼠标x坐标"
    if opcode == "sensing_mousey":
        return "鼠标y坐标"

    # 菜单和普通输入值
    for field_name in [
        "COSTUME", "BACKDROP", "BROADCAST_OPTION", "VARIABLE", "KEY_OPTION",
        "TO", "TOWARDS", "SOUND_MENU", "CLONE_OPTION"
    ]:
        value = _scratch_field(block, field_name)
        if value:
            return value
    return opcode or "未识别表达式"


def _scratch_block_line(blocks: dict, block_id: str) -> str:
    block = blocks.get(block_id, {})
    if not isinstance(block, dict):
        return ""
    op = block.get("opcode", "")
    inp = lambda name: _scratch_input(blocks, block, name, seen={block_id})
    fld = lambda name: _scratch_field(block, name)

    mapping = {
        "event_whenflagclicked": "当点击绿旗",
        "event_whenthisspriteclicked": "当角色被点击",
        "control_forever": "循环执行",
        "control_stop": f"停止{fld('STOP_OPTION') or '全部'}",
        "motion_ifonedgebounce": "碰到边缘就反弹",
        "looks_show": "显示",
        "looks_hide": "隐藏",
        "pen_clear": "全部擦除",
        "pen_penDown": "落笔",
        "pen_penUp": "抬笔",
    }
    if op in mapping:
        return mapping[op]
    if op == "event_whenbroadcastreceived":
        return f"当接收到{fld('BROADCAST_OPTION')}"
    if op == "event_whenkeypressed":
        return f"当按下{fld('KEY_OPTION')}键"
    if op == "event_broadcast":
        return f"广播{inp('BROADCAST_INPUT')}"
    if op == "event_broadcastandwait":
        return f"广播{inp('BROADCAST_INPUT')}并等待"
    if op == "control_wait":
        return f"等待{inp('DURATION')}秒"
    if op == "control_repeat":
        return f"重复执行{inp('TIMES')}次"
    if op == "control_repeat_until":
        return f"重复执行直到{inp('CONDITION')}"
    if op == "control_wait_until":
        return f"等待直到{inp('CONDITION')}"
    if op == "control_if":
        return f"如果{inp('CONDITION')}那么"
    if op == "control_if_else":
        return f"如果{inp('CONDITION')}那么/否则"
    if op == "looks_switchcostumeto":
        return f"换成{inp('COSTUME')}造型"
    if op == "looks_nextcostume":
        return "下一个造型"
    if op == "looks_sayforsecs":
        return f"说{inp('MESSAGE')}{inp('SECS')}秒"
    if op == "looks_say":
        return f"说{inp('MESSAGE')}"
    if op == "looks_setsizeto":
        return f"将大小设为{inp('SIZE')}%"
    if op == "looks_changesizeby":
        return f"将大小增加{inp('CHANGE')}"
    if op == "motion_goto":
        return f"移到{inp('TO')}"
    if op == "motion_gotoxy":
        return f"移到x:{inp('X')} y:{inp('Y')}"
    if op == "motion_glidesecstoxy":
        return f"在{inp('SECS')}秒内滑行到x:{inp('X')} y:{inp('Y')}"
    if op == "motion_movesteps":
        return f"移动{inp('STEPS')}步"
    if op == "motion_changexby":
        return f"将x坐标增加{inp('DX')}"
    if op == "motion_changeyby":
        return f"将y坐标增加{inp('DY')}"
    if op == "motion_setx":
        return f"将x坐标设为{inp('X')}"
    if op == "motion_sety":
        return f"将y坐标设为{inp('Y')}"
    if op == "motion_pointindirection":
        return f"面向{inp('DIRECTION')}方向"
    if op == "motion_pointtowards":
        return f"面向{inp('TOWARDS')}"
    if op == "motion_turnleft":
        return f"左转{inp('DEGREES')}度"
    if op == "motion_turnright":
        return f"右转{inp('DEGREES')}度"
    if op == "motion_setrotationstyle":
        return f"将旋转方式设为{fld('STYLE')}"
    if op == "data_setvariableto":
        return f"将{fld('VARIABLE')}设为{inp('VALUE')}"
    if op == "data_changevariableby":
        return f"将{fld('VARIABLE')}增加{inp('VALUE')}"
    if op == "sound_playuntildone":
        return f"播放声音{inp('SOUND_MENU')}等待播完"
    if op == "sound_play":
        return f"播放声音{inp('SOUND_MENU')}"
    if op == "pen_setPenColorToColor":
        return f"将画笔颜色设为{inp('COLOR')}"
    if op == "pen_setPenSizeTo":
        return f"将画笔粗细设为{inp('SIZE')}"
    return op or "未识别积木"


def _scratch_render_chain(blocks: dict, start_id: str, depth=0, seen=None, max_blocks=120):
    """按真实连接关系读取一段脚本，并保留循环/条件内部的层级。"""
    seen = set(seen or set())
    lines = []
    current = start_id
    count = 0
    while current and current not in seen and count < max_blocks:
        seen.add(current)
        count += 1
        block = blocks.get(current, {})
        if not isinstance(block, dict):
            break
        line = _scratch_block_line(blocks, current)
        if line:
            lines.append("    " * depth + line)
        opcode = block.get("opcode", "")
        if opcode in {"control_forever", "control_repeat", "control_repeat_until", "control_if", "control_if_else"}:
            sub = block.get("inputs", {}).get("SUBSTACK")
            sub_id = sub[1] if isinstance(sub, list) and len(sub) > 1 and isinstance(sub[1], str) else None
            if sub_id:
                lines.extend(_scratch_render_chain(blocks, sub_id, depth + 1, seen, max_blocks))
            if opcode == "control_if_else":
                sub2 = block.get("inputs", {}).get("SUBSTACK2")
                sub2_id = sub2[1] if isinstance(sub2, list) and len(sub2) > 1 and isinstance(sub2[1], str) else None
                if sub2_id:
                    lines.append("    " * depth + "否则")
                    lines.extend(_scratch_render_chain(blocks, sub2_id, depth + 1, seen, max_blocks))
        current = block.get("next")
    return lines


@st.cache_data(show_spinner=False)
def load_teacher_theme_reference(theme: str, file_path: str, modified_time: float) -> dict:
    """读取教师基础版 .sb3，提取每个角色的真实脚本，供后台理解任务目标。"""
    _ = modified_time
    result = {"theme": theme, "file": os.path.basename(file_path), "roles": [], "errors": []}
    try:
        with zipfile.ZipFile(file_path, "r") as archive:
            project = json.loads(archive.read("project.json").decode("utf-8"))
        for target in project.get("targets", []):
            blocks = target.get("blocks", {}) or {}
            top_ids = [
                block_id for block_id, block in blocks.items()
                if isinstance(block, dict) and safe_bool(block.get("topLevel", False))
            ]
            top_ids.sort(key=lambda block_id: (
                float((blocks.get(block_id, {}) or {}).get("y", 0) or 0),
                float((blocks.get(block_id, {}) or {}).get("x", 0) or 0),
            ))
            scripts = []
            for block_id in top_ids:
                lines = _scratch_render_chain(blocks, block_id)
                if lines:
                    scripts.append(lines)
            role = {
                "name": str(target.get("name", "") or ""),
                "is_stage": safe_bool(target.get("isStage", False)),
                "costumes": [str(item.get("name", "") or "") for item in target.get("costumes", []) if isinstance(item, dict)],
                "variables": [str(value[0]) for value in (target.get("variables", {}) or {}).values() if isinstance(value, list) and value],
                "scripts": scripts,
            }
            if scripts or role["costumes"]:
                result["roles"].append(role)
    except Exception as exc:
        result["errors"].append(f"教师基础版读取失败：{exc}")
    return result


def get_teacher_theme_reference(theme: str) -> dict:
    file_name = THEME_REFERENCE_SB3_FILES.get(str(theme or "").strip(), "")
    if not file_name:
        return {"theme": theme, "roles": [], "errors": ["当前主题未配置教师基础版参照文件"]}
    path = APP_DIR / file_name
    if not path.exists():
        return {"theme": theme, "roles": [], "errors": [f"未找到教师基础版文件：{file_name}"]}
    return load_teacher_theme_reference(theme, str(path), path.stat().st_mtime)


def _reference_role_score(role: dict, current_object: str, question: str) -> int:
    name = re.sub(r"\s+", "", str(role.get("name", "") or ""))
    obj = re.sub(r"\s+", "", str(current_object or ""))
    q = re.sub(r"\s+", "", str(question or ""))
    score = 0
    if name and obj and (name == obj or name in obj or obj in name):
        score += 20
    if name and name in q:
        score += 10
    role_text = "".join("".join(script) for script in role.get("scripts", []))
    for keyword in ["哭", "笑", "锤子", "苹果", "老鼠", "小猫", "螃蟹", "鱼", "得分", "广播"]:
        if keyword in q and keyword in (name + role_text):
            score += 3
    if role.get("is_stage"):
        score -= 2
    return score


def build_theme_reference_context(theme: str, current_object: str = "", student_question: str = "") -> str:
    """
    生成后台主题参照。教师基础版只用于理解基础功能目标，
    不是要求学生逐块照抄的唯一标准答案。
    """
    reference = get_teacher_theme_reference(theme)
    if reference.get("errors") and not reference.get("roles"):
        return "教师基础版主题参照暂不可用：" + "；".join(reference.get("errors", []))

    roles = list(reference.get("roles", []))
    roles.sort(key=lambda role: _reference_role_score(role, current_object, student_question), reverse=True)
    selected = [role for role in roles if _reference_role_score(role, current_object, student_question) > 0][:4]
    if not selected:
        selected = [role for role in roles if not role.get("is_stage")][:4]

    lines = [
        f"主题：{theme}",
        f"教师基础版文件：{reference.get('file', '')}",
        "用途：理解该主题基础功能、角色关系和正常运行逻辑；不得把教师代码当作学生必须逐块照抄的唯一答案。",
    ]
    for role in selected:
        lines.append(f"\n角色：{role.get('name', '')}")
        if role.get("costumes"):
            lines.append("造型：" + "、".join(role.get("costumes", [])))
        for index, script in enumerate(role.get("scripts", [])[:6], start=1):
            lines.append(f"基础脚本{index}：")
            lines.extend(script[:40])
    if reference.get("errors"):
        lines.append("参照读取提示：" + "；".join(reference.get("errors", [])))
    return "\n".join(lines)[:12000]


def classify_student_answer_mode(question: str, has_screenshot: bool = False) -> str:
    text = re.sub(r"\s+", "", str(question or ""))
    if any(word in text for word in ["创新", "优化", "拓展", "扩展", "升级", "更有趣", "还能加", "美化", "创意"]):
        return "innovation"
    if has_screenshot:
        return "screenshot_diagnosis"
    if any(word in text for word in ["怎么做", "如何做", "怎样完成", "不会做", "从哪里开始"]):
        return "basic_guidance"
    return "normal"


def build_theme_mode_prompt(mode: str, theme_reference_context: str) -> str:
    return f"""
【教师基础版主题参照】
{theme_reference_context or '当前没有可用的教师基础版主题参照。'}

【当前回答模式】
{mode}

【模式规则】
1. screenshot_diagnosis：只诊断学生当前截图和问题。先判断学生想实现的效果，再对照截图与教师基础版目标，选择证据最充分的一个相关问题，只给一个具体改法；不得创新、不得删除正常基础功能、不得把教师基础版当作唯一代码写法。
2. basic_guidance：先帮助学生完成教师基础版要求的核心功能，不主动加入计分、关卡、音效等拓展功能。
3. innovation：只有学生明确询问创新、优化或拓展时，才在已经完成基础功能的前提下给1—2个可行创意；不得把创新建议混入普通调试回答。
4. normal：结合当前主题回答，但不强行套用教师基础版代码。
5. 教师基础版用于理解“应该实现什么效果”，不是用来要求学生代码逐块完全一致。
""".strip()


def _parse_json_object(raw_text: str) -> dict:
    text = str(raw_text or "").strip()
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.I)
    text = re.sub(r"\s*```$", "", text)
    try:
        data = json.loads(text)
        return data if isinstance(data, dict) else {}
    except Exception:
        start, end = text.find("{"), text.rfind("}")
        if start >= 0 and end > start:
            try:
                data = json.loads(text[start:end + 1])
                return data if isinstance(data, dict) else {}
            except Exception:
                return {}
    return {}


def diagnose_screenshot_with_teacher_reference(
    student_question: str,
    screenshot_structure: dict,
    ocr_text: str,
    theme_reference_context: str,
    answer_mode: str,
) -> dict:
    """第二阶段诊断：以截图事实和教师基础目标为依据，输出一个可校验的诊断 JSON。"""
    prompt = f"""
你是小学五年级图形化编程课堂的程序调试诊断器。你不能自由发挥，也不能为了显得完整而猜测。

【学生问题】
{student_question or '学生只上传截图，请找出一个最明确的问题。'}

【当前模式】
{answer_mode}

【学生截图结构】
{json.dumps(screenshot_structure, ensure_ascii=False, indent=2)}

【OCR文字，仅用于核对积木原文】
{str(ocr_text or '')[:2500]}

【教师基础版主题参照】
{theme_reference_context}

请先在内部完成：
1. 明确学生想实现的触发方式和结果；
2. 找到与问题直接相关的角色和脚本；
3. 逐段读取所有可见脚本，但只选择一个与学生问题最相关、证据最充分的问题；
4. 教师基础版只用于理解基础功能目标，不要求学生逐块照抄；
5. 证据不足时必须追问，不能猜测脚本冲突、删除脚本、添加变量或重写整个程序；
6. 如果需要给出坐标、角度、秒数、次数、变量值等具体数值，必须同时说明依据；例如“从 x=-98 开始，后面增加 130，所以终点是 32”；
7. 普通调试不得提供创新功能。只有 mode=innovation 时才给创意。

严格只输出以下 JSON，不要输出解释或代码围栏：
{{
  "status": "answer|clarify",
  "mode": "screenshot_diagnosis|innovation|basic_guidance|normal",
  "relevant_role": "截图中相关角色原名",
  "student_goal": "一句话说明学生想实现的效果",
  "confirmed_evidence": ["截图中直接可见的事实，最多3条"],
  "primary_reason": "用自然语言指出截图里和学生问题最相关的设置或现象；证据不足时留空",
  "change_instruction": "只修改一个关键点，明确说把什么改成什么或移动到哪里；如果给出具体数值，要写出这个数值的来源或计算；证据不足时留空",
  "program_lines": ["仅在连接关系和原文都足够确定时给局部程序；否则为空数组"],
  "program_grounded": false,
  "verification": "修改后让学生观察什么",
  "clarifying_question": "status=clarify时只问一个问题",
  "confidence": 0.0
}}
"""
    try:
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=1300,
            stream=False,
        )
        data = _parse_json_object(response.choices[0].message.content or "")
    except Exception as exc:
        return {"status": "clarify", "clarifying_question": f"截图分析暂时失败，请再发送一次清晰的局部截图。", "confidence": 0.0, "error": str(exc)}

    status = str(data.get("status", "clarify") or "clarify")
    confidence = safe_float(data.get("confidence", 0.0), 0.0)
    evidence = data.get("confirmed_evidence", [])
    if not isinstance(evidence, list):
        evidence = []
    data["confirmed_evidence"] = [str(item) for item in evidence if str(item).strip()][:3]
    lines = data.get("program_lines", [])
    if not isinstance(lines, list):
        lines = []
    data["program_lines"] = [str(item) for item in lines if str(item).strip()][:14]
    data["confidence"] = confidence

    # 低置信度、没有截图证据或没有明确修改时，统一转为追问，避免强行给错答案。
    if answer_mode != "innovation":
        if status != "answer" or confidence < 0.72 or not data.get("primary_reason") or not data.get("change_instruction"):
            data["status"] = "clarify"
            if not data.get("clarifying_question"):
                data["clarifying_question"] = "从这张截图还不能确定问题。运行时具体出现了什么现象？"
            data["program_lines"] = []
            data["program_grounded"] = False
    return data


def render_grounded_student_answer(diagnosis: dict) -> str:
    if str(diagnosis.get("status", "")) != "answer":
        return str(diagnosis.get("clarifying_question", "") or "从这张截图还不能确定问题，请补充运行现象。")

    if diagnosis.get("mode") == "innovation":
        reason = str(diagnosis.get("primary_reason", "") or "")
        change = str(diagnosis.get("change_instruction", "") or "")
        return "\n".join(item for item in [reason, change] if item).strip()

    parts = []
    reason = str(diagnosis.get("primary_reason", "") or "").strip()
    change = str(diagnosis.get("change_instruction", "") or "").strip()
    if reason:
        parts.append(reason)
    if change:
        parts.append(change)

    program_lines = diagnosis.get("program_lines", []) or []
    if safe_bool(diagnosis.get("program_grounded", False)) and safe_float(diagnosis.get("confidence", 0), 0) >= 0.84 and program_lines:
        parts.append("修改后像这样：\n[[PROGRAM]]\n" + "\n".join(program_lines) + "\n[[/PROGRAM]]")

    verification = str(diagnosis.get("verification", "") or "").strip()
    if verification:
        parts.append(verification)
    elif change:
        parts.append("改好后运行一下，看看这个现象有没有变化。")
    return "\n\n".join(parts).strip()


# =========================
# 1.5 页面样式
# =========================

def inject_custom_css():
    st.markdown(
        """
        <style>
        .stApp {
            background:
                radial-gradient(circle at 14% 12%, rgba(255, 210, 117, 0.18) 0, rgba(255, 210, 117, 0) 26%),
                radial-gradient(circle at 86% 8%, rgba(112, 169, 255, 0.16) 0, rgba(112, 169, 255, 0) 28%),
                linear-gradient(180deg, #F7FBFF 0%, #EEF6FF 48%, #F8FBFF 100%);
        }

        .block-container {
            max-width: 1240px;
            padding-top: 2.6rem;
            padding-bottom: 7.4rem;
        }

        section[data-testid="stSidebar"] {
            background: #FFFFFF;
            border-right: 1px solid #E6EDF5;
        }

        button[data-testid="stSidebarCollapseButton"],
        div[data-testid="stSidebarCollapseButton"] {
            display: none !important;
        }

        #MainMenu, footer {
            visibility: hidden;
        }

        .main-card {
            background: rgba(255, 255, 255, 0.92);
            border: 1px solid #E3ECF6;
            border-radius: 20px;
            padding: 17px 22px;
            box-shadow: 0 12px 28px rgba(52, 86, 130, 0.08);
            margin-bottom: 16px;
        }

        .title-row {
            display: flex;
            align-items: center;
            gap: 14px;
        }

        .title-icon {
            width: 48px;
            height: 48px;
            border-radius: 15px;
            background: linear-gradient(135deg, #FFF2D6 0%, #EAF4FF 100%);
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 1.65rem;
            box-shadow: inset 0 0 0 1px rgba(224, 235, 248, 0.9);
            flex: 0 0 48px;
        }

        .title-text {
            min-width: 0;
        }

        .app-title {
            font-size: 1.82rem;
            line-height: 1.16;
            font-weight: 850;
            color: #273247;
            margin: 0 0 5px 0;
            letter-spacing: -0.02em;
        }

        .app-desc {
            font-size: 0.92rem;
            color: #66758F;
            margin: 0;
        }

        div[data-testid="stTextInput"] input {
            border-radius: 14px;
            border: none !important;
            background: #FFFFFF;
            height: 42px;
            box-shadow: none !important;
            outline: none !important;
        }

        div[data-testid="stTextInput"] [data-baseweb="input"] {
            border-radius: 14px !important;
            border: 1px solid #DDE8F4 !important;
            background: #FFFFFF !important;
            box-shadow: none !important;
        }

        div[data-testid="stTextInput"] [data-baseweb="input"]:focus-within {
            border: 1px solid #80B7FF !important;
            box-shadow: 0 0 0 2px rgba(128, 183, 255, 0.12) !important;
        }

        div[data-testid="stTextInput"] input:focus {
            outline: none !important;
            box-shadow: none !important;
            border: none !important;
        }

        div[data-testid="stTextInput"] label {
            color: #263449;
            font-weight: 600;
        }

        .starter-area {
            margin: 0 0 8px 38px;
            max-width: 500px;
        }

        .starter-area .stButton {
            margin-bottom: 1px;
        }

        .starter-area .stButton > button {
            width: fit-content;
            max-width: 100%;
            min-height: 22px;
            border-radius: 9px;
            border-top-left-radius: 4px;
            border: 1px solid #E5EDF6;
            background: rgba(255, 255, 255, 0.88);
            color: #44536A;
            padding: 0.12rem 0.36rem;
            text-align: left;
            font-size: 0.68rem;
            line-height: 1.2;
            box-shadow: 0 2px 6px rgba(52, 86, 130, 0.025);
        }

        .starter-area .stButton > button:hover {
            border-color: #9BC7FF;
            color: #1F5FBF;
            background: #FFFFFF;
            transform: translateY(-1px);
        }

        .stButton > button {
            border-radius: 15px;
            border: 1px solid #D8E5F4;
            background: rgba(255,255,255,0.94);
            color: #3E5069;
            padding: 0.48rem 0.74rem;
            min-height: 40px;
            font-size: 0.88rem;
            box-shadow: 0 7px 16px rgba(52, 86, 130, 0.055);
            transition: all 0.16s ease-in-out;
            text-align: left;
        }

        .stButton > button:hover {
            border-color: #8EBEFF;
            color: #1F5FBF;
            background: #FFFFFF;
            transform: translateY(-1px);
        }

        .chat-wrap {
            display: flex;
            width: 100%;
            margin: 10px 0;
            align-items: flex-start;
            gap: 8px;
        }

        .chat-wrap.user {
            justify-content: flex-end;
        }

        .chat-wrap.assistant {
            justify-content: flex-start;
        }

        .chat-content {
            min-width: 0;
            display: flex;
        }

        .chat-wrap.user .chat-content {
            max-width: 78%;
            justify-content: flex-end;
        }

        .chat-wrap.assistant .chat-content {
            max-width: 88%;
            justify-content: flex-start;
        }

        .chat-bubble {
            display: inline-block;
            width: fit-content;
            max-width: 100%;
            padding: 10px 13px;
            border-radius: 16px;
            font-size: 0.95rem;
            line-height: 1.55;
            white-space: pre-wrap;
            word-break: break-word;
            box-sizing: border-box;
            box-shadow: 0 6px 18px rgba(52, 86, 130, 0.06);
        }

        .chat-bubble .chat-image {
            display: block;
            max-width: 260px;
            max-height: 190px;
            border-radius: 12px;
            margin-top: 8px;
            border: 1px solid #E5EDF6;
            object-fit: contain;
            background: #FFFFFF;
        }

        .pending-image-box {
            display: inline-flex;
            align-items: center;
            gap: 10px;
            background: rgba(255,255,255,0.92);
            border: 1px solid #DDE8F4;
            border-radius: 14px;
            padding: 8px 10px;
            margin: 6px 0 8px 0;
            box-shadow: 0 6px 16px rgba(52, 86, 130, 0.05);
            color: #526277;
            font-size: 0.86rem;
        }

        .chat-bubble.user {
            background: #B9ECA0;
            color: #1F2B1D;
            border-top-right-radius: 5px;
        }

        .chat-bubble.assistant {
            background: #FFFFFF;
            color: #273247;
            border: 1px solid #E5EDF6;
            border-top-left-radius: 5px;
        }

        .chat-bubble .md-line {
            margin: 0.24rem 0;
        }

        .chat-bubble .md-heading {
            font-weight: 850;
            font-size: 1.08rem;
            color: #21304A;
            margin: 0.72rem 0 0.36rem 0;
        }

        .chat-bubble .md-heading:first-child {
            margin-top: 0;
        }

        .chat-bubble .md-number,
        .chat-bubble .md-bullet {
            margin: 0.28rem 0;
        }

        .chat-bubble .md-space {
            height: 0.45rem;
        }

        .chat-bubble .md-table-wrap {
            width: 100%;
            overflow-x: auto;
            margin: 0.45rem 0;
        }

        .chat-bubble table.md-table {
            border-collapse: collapse;
            width: 100%;
            min-width: 760px;
            font-size: 0.88rem;
            line-height: 1.45;
            background: #FFFFFF;
        }

        .chat-bubble table.md-table th,
        .chat-bubble table.md-table td {
            border: 1px solid #DDE8F4;
            padding: 8px 9px;
            text-align: left;
            vertical-align: top;
            white-space: normal;
        }

        .chat-bubble table.md-table th {
            background: #F3F8FF;
            font-weight: 800;
            color: #21304A;
        }

        .chat-bubble .program-block {
            margin: 0.55rem 0;
            border-left: 4px solid #D9DEE7;
            background: #F7F8FA;
            border-radius: 8px;
            padding: 10px 12px;
            overflow-x: auto;
        }

        .chat-bubble .program-block pre {
            margin: 0;
            white-space: pre-wrap;
            word-break: break-word;
            font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
            font-size: 0.92rem;
            line-height: 1.65;
            color: #202939;
        }

        .chat-bubble strong {
            color: #1F2B3F;
            font-weight: 850;
        }

        .chat-avatar {
            width: 30px;
            height: 30px;
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 0.9rem;
            flex: 0 0 30px;
            margin-top: 2px;
            box-shadow: 0 4px 12px rgba(52, 86, 130, 0.08);
        }

        .chat-avatar.user {
            background: #DFF6D3;
            color: #2F6B27;
            order: 2;
        }

        .chat-avatar.assistant {
            background: #FFFFFF;
            color: #273247;
            border: 1px solid #E5EDF6;
        }

        div[data-testid="stChatInput"] {
            bottom: 10px !important;
            background: transparent !important;
            padding-bottom: 0 !important;
        }

        div[data-testid="stChatInput"] > div {
            border: 1px solid #CFE0F5 !important;
            box-shadow: 0 8px 22px rgba(52, 86, 130, 0.07) !important;
        }

        div[data-testid="stChatInput"]:focus-within > div {
            border: 1px solid #80B7FF !important;
            box-shadow: 0 0 0 2px rgba(128, 183, 255, 0.16) !important;
        }

        div[data-testid="stChatInput"] [data-baseweb="textarea"],
        div[data-testid="stChatInput"] textarea {
            border: none !important;
            box-shadow: none !important;
            outline: none !important;
            background: transparent !important;
        }

        div[data-testid="stChatInput"] [data-baseweb="textarea"]:focus-within,
        div[data-testid="stChatInput"] textarea:focus {
            border: none !important;
            box-shadow: none !important;
            outline: none !important;
        }

        /* 底部截图工具条：放在输入框下方，不压住输入框，不进入侧边栏 */
        .st-key-paste_toolbar {
            position: fixed;
            left: max(350px, calc((100vw - 1240px) / 2 + 2.2rem));
            right: max(2rem, calc((100vw - 1240px) / 2 + 1rem));
            bottom: 4px !important;
            z-index: 999999;
            background: transparent !important;
            border: none !important;
            box-shadow: none !important;
            padding: 0 !important;
            min-height: 0 !important;
            height: auto !important;
            overflow: visible;
            pointer-events: auto;
        }

        .st-key-paste_toolbar [data-testid="stHorizontalBlock"] {
            align-items: center;
            gap: 0.28rem;
        }
        .st-key-paste_toolbar .stButton > button {
            min-height: 26px !important;
            padding: 0.12rem 0.42rem !important;
            border-radius: 11px;
            font-size: 0.78rem;
            text-align: center;
            white-space: nowrap;
            box-shadow: none;
        }

        .st-key-paste_toolbar .pending-image-box {
            margin: 0;
            padding: 3px 8px;
            font-size: 0.76rem;
            white-space: nowrap;
            box-shadow: none;
        }

        .st-key-paste_toolbar img {
            max-height: 24px !important;
            width: auto !important;
            border-radius: 7px;
            border: 1px solid #E5EDF6;
            object-fit: contain;
        }

        </style>
        """,
        unsafe_allow_html=True
    )


inject_custom_css()


# =========================
# 2. 系统提示词
# =========================

SYSTEM_PROMPT = """
你是“图形化编程学习助手”，服务于小学高年级图形化编程课堂。

你的服务对象包括教师和学生。

总原则：
1. 教师端重在支持教学设计、课堂调控、任务单生成、问题链设计、调试提示、展示评价和教学反思。
2. 学生端重在引导思考，不能直接替学生完成完整程序，不能直接给出完整答案。
3. 所有回答必须围绕小学图形化编程任务展开。
4. 面向学生时，要使用自然、简短、适合小学五年级学生理解的语言；问题简单时直接回答，确实存在多个步骤时再分步提示。
5. 面向教师时，要规范、清晰、可操作，贴近小学课堂实际。
6.图形化编程教学流程为：情境创设、任务分析、实践创作、展示评价。
7. 任务分析建议采用：作品 → 角色 → 动作、规则、效果；必要时可引导教师和学生借助思维导图、流程图或任务分析表梳理作品结构与程序逻辑。
8. 学生调试时优先提示检查：启动事件、角色脚本归属、重复执行、条件是否满足、变量是否初始化、角色是否隐藏、造型、位置、方向是否设置正确。
9. 智能体始终遵循：教师主导，学生主体，智能辅助。
10. 不生成脱离图形化编程课堂实际的空泛内容。
11. 回答中不得出现 <br>、<p>、<div> 等 HTML 标签。
12. 学生端不使用 Markdown 表格。教师端生成教学设计时，“教学过程”必须使用 Markdown 表格呈现，表头固定为：教学环节、教师活动、学生活动、智能体支持、设计意图。
13. 教师端除教学过程表格外，其他部分可采用“小标题 + 分点说明”的形式呈现。
14. 学生端分条使用“①②③”。教师端注意编号层级：一级标题用（一）（二）（三），二级条目用1. 2. 3.，三级细目用①②③，四级细目用（1）（2）（3），不要同一层级混用。
15. 不要输出 HTML 标签，不要输出网页代码。
16. 不要输出空编号、空项目或占位符；如果信息不确定，应直接追问学生补充，而不是留下空白。
17. 不要使用 Markdown 加粗或斜体符号，例如 **重点**、*内容*、`代码`。
18. 生成教学设计时必须保证每个环节的“教师活动、学生活动、智能体支持、设计意图”完整输出；如果内容较长，应压缩前文，不得省略后文或停在“通过……”这类未完成句子。
"""


def build_role_prompt(role: str) -> str:
    if role == "教师端":
        return """
当前用户身份：教师。

你是小学高年级图形化编程教学支持助手。

请根据教师输入，自动判断其需求属于以下哪一类：
1. 教学设计；
2. 任务单设计；
3. 任务分析模板；
4. 课堂问题链；
5. 调试提示；
6. 展示评价；
7. 教学反思；
8. 智能体使用设计。

回答要求：

一、总体要求
1. 内容必须贴近小学高年级图形化编程课堂。
2. 不写空泛套话，要具体、清晰、可操作。
3. 教学流程优先采用：情境创设 → 任务分析 → 实践创作 → 展示评价。
4. 任务分析统一采用“主题 → 背景 → 角色”的方式，其中角色部分重点分析动作、规则和效果；如果作品没有背景，不要强行补充；在任务分析环节应鼓励教师使用思维导图、流程图或任务分析表帮助学生梳理作品结构、角色关系和程序逻辑。
5. 除“教学过程”外，一般不使用 Markdown 表格；“教学过程”必须使用 Markdown 表格呈现。
6. 不得出现 <br>、<p>、<div> 等 HTML 标签。
7. 教学过程表格必须包含五列：教学环节、教师活动、学生活动、智能体支持、设计意图；每个环节分别写清对应内容。
9. 生成教学设计时必须保证每个环节的“设计意图”完整输出；如果内容较长，应适当压缩前文，不要省略后文。
8. 语言要规范，适合教师直接修改后用于论文、教案或课堂材料。

二、当教师要求生成教学设计时，必须按照以下结构输出，并注意编号层级：一级标题用（一）（二）（三），二级条目用1. 2. 3.，三级细目用①②③，四级细目用（1）（2）（3），不要同一层级混用：

（一）教学设计基础

必须包括以下内容：

1. 教学对象
说明适用年级、学生已有基础和学习背景。

2. 学情分析
从以下三个方面展开：
① 认知基础：学生对角色、舞台、积木、程序逻辑等已有理解。
② 技能基础：学生是否具备基本拖拽积木、运行程序、修改脚本等操作经验。
③ 学习特点：学生可能存在的兴趣特点、操作差异、任务理解困难或调试困难。

3. 教学目标
按照义务教育信息科技课程标准中的四个核心素养维度表述：
① 信息意识；
② 计算思维；
③ 数字化学习与创新；
④ 信息社会责任。

目标要结合具体图形化编程作品主题，不要空泛。

4. 教学重难点
① 教学重点：本节课学生必须掌握的核心知识、关键积木或程序结构。
② 教学难点：学生在任务分析、程序搭建或调试优化中可能遇到的主要困难。

5. 教学策略
重点说明突破重难点的策略，例如：
① 通过作品展示引出任务，激发学生创作兴趣；
② 采用“作品 → 角色 → 动作、规则、效果”的方式进行任务拆解；
③ 教师通过板书、投屏或任务单提供文字版任务分析框架，引导学生梳理作品结构、角色关系和程序逻辑；
④ 学生在思路不清时，可借助智能体整理作品逻辑、生成文字版任务分析框架或完善任务分析表，但最终分析结果应由学生自行修改和完善；
⑤ 通过教师关键示范和学生实践结合促进理解；
⑥ 通过巡视指导、同伴交流和共性问题讲解帮助学生调试修改。

6. 教学环境与资源
包括图形化编程平台、计算机或平板设备、作品素材、任务单、思维导图、流程图、评价表、智能体等。

（二）教学过程

教学过程必须使用 Markdown 表格输出，表头固定为：
| 教学环节 | 教师活动 | 学生活动 | 智能体支持 | 设计意图 |
| --- | --- | --- | --- | --- |

表格至少包含以下四个环节：情境创设、任务分析、实践创作、展示评价。每个单元格内容要具体、完整、可操作，不要只写短语。

特别注意：
1. 教学过程表格必须严格使用 Markdown 表格。
2. 每一行必须以 | 开头并以 | 结尾。
3. 每个教学环节只能占一整行，不能在单元格内部换行。
4. 单元格内不要使用 ①②③、项目符号或换行分点，可以用分号连接多个活动。
5. 不要在表格中插入空行。
6. 表格行格式示例：
| 情境创设 | 教师播放作品并提出问题，引导学生观察角色和规则 | 学生观看作品，回答角色、动作和得分方式 | 本环节不直接使用智能体，教师可课前借助智能体生成导入问题 | 激发兴趣，明确学习任务 |


（三）教学反思

教学反思必须作为教学设计结尾，简要包括以下内容：

1. 目标达成反思：说明本节课教学目标是否基本达成。
2. 学生学习反思：说明学生在任务理解、程序搭建、调试修改中的表现。
3. 智能体使用反思：说明智能体在哪些环节提供了支持，是否存在使用不足。
4. 改进建议：提出后续教学中可以优化的地方。

三、关于“智能体支持”的要求

1. 智能体支持必须根据具体课的内容和环节合理填写，不得机械套用。
2. 情境创设环节一般不安排学生直接使用智能体。可以写“教师课前借助智能体生成导入问题或旧知唤醒问题”，也可以写“本环节不直接使用智能体”。
3. 任务分析环节可以体现教师使用思维导图引导全班分析任务，学生也可以向智能体描述自己的作品想法，由智能体帮助梳理“作品 → 角色 → 动作、规则、效果”、角色关系、功能模块和程序运行逻辑，并形成可用于完善思维导图或任务分析表的提示。
4. 实践创作环节是智能体支持的重点，可以体现学生在程序调试、积木提示、错误排查、作品优化中向智能体获取分步提示。
5. 展示评价环节不得写“智能体对作品进行评分”。智能体不能替代教师进行正式评价。
6. 展示评价环节可以写“教师课前借助智能体生成评价问题”或“学生可借助智能体整理展示表达思路”，但正式评价仍由教师依据评价量表完成。
7. 如果某一环节不适合使用智能体，可以明确写“本环节不直接使用智能体”。

四、当教师要求生成任务单时

尽量包含：
1. 作品名称；
2. 任务分析；
3. 基础任务；
4. 提升任务；
5. 自我检查。

任务单也不要使用 Markdown 表格，可以用标题和分点呈现。

五、当教师要求生成调试提示时

要适合教师课堂讲解和巡视指导，围绕学生常见问题展开，例如：
1. 角色不动；
2. 角色只动一次；
3. 分数不增加；
4. 变量没有初始化；
5. 程序写错角色；
6. 条件判断没有触发；
7. 角色隐藏后没有显示；
8. 碰到边缘后角色方向异常。

六、当教师要求生成评价建议时

可以围绕完整性、技术性、创新性和艺术性展开，但不得让智能体替代教师评分。

七、如果涉及实验班和对照班

需说明：
1. 实验班使用智能体支持；
2. 对照班采用常规教学支持；
3. 两班在教学主题、课时安排、学习任务和作品要求上保持一致。
"""
    else:
        return """
当前用户身份：学生。

你是小学图形化编程学习助手。你的回答应像教师在学生身边进行即时指导，既要准确，也要自然、简短，避免使用固定模板和重复句式。
请根据学生输入，自动判断其问题属于以下哪一类：
1. 任务理解；
2. 任务分析；
3. 积木功能理解；
4. 程序实现提示；
5. 程序调试；
6. 作品优化；
7. 展示表达。

学生端回答规则：

一、基本规则
1. 不能直接替学生完成整个作品或给出所有角色的完整程序。对于学生已经搭建并上传截图的调试问题，在准确说明原因后，可以展示“与当前错误直接相关的一段修改后程序”，帮助学生对照修改，但不得扩展为整个作品答案。
2. 回答应使用小学五年级学生能够理解的短句，避免长篇解释和过多专业术语。
3. 学生的问题已经明确时，直接回应核心问题，不重复复述学生的问题。
3.1 当前正在和学生本人对话，回答必须直接使用“你”“你的程序”“你把……”，不要用“学生”“该学生”“学生把……”等第三人称称呼对方。只有描述一般教学规则时才可以使用“学生”一词。
4. 同一段连续对话中，不要每次都使用“同学你好”“我看到你的问题了”“我看到你上传的截图了”等问候或过渡语。首次交流时可以自然问候，后续应直接进入问题。
5. 回答不使用固定模板。可以先在内部判断问题、原因和修改方向，但输出时要组织成自然的课堂对话，不必机械地分成“问题、原因、建议”三个部分。
6. 内容较少时，可以直接用一段话回答；只有确实存在多个步骤、角色或检查项目时，才使用①②③分点。
7. 分点提示统一使用“①②③”，不要使用“（1）（2）（3）”或“1. 2. 3.”。
8. 如果信息不足，最多追问2个关键问题。调试类问题通常只追问1个最关键的问题。
9. 如果是调试问题，优先判断一个最可能的原因，并给出一个最关键的检查或修改方向。
10. 如果当前信息只能确定检查方向，不能把推测当作已经确认的错误，应使用“先检查……”“看看是否……”等谨慎表达。
11. 如果是任务分析，采用“作品 → 角色 → 动作、规则、效果”的方式引导。
12. 如果学生说“帮我画思维导图”“帮我梳理逻辑”“我不知道怎么分析作品”，应帮助学生整理文字版任务分析框架，但不能替学生完成全部创作内容。
13. 如果是作品优化，只给1—2个与当前作品最相关、学生能够自己尝试的建议。
14. 鼓励学生修改后运行并观察效果，但不要每次都固定使用“再运行看看”“试试看”等相同结尾。
15. 不使用Markdown表格，不得出现<br>、<p>、<div>等HTML标签。
16. 不要使用Markdown加粗、斜体或代码符号，例如**重点**、*内容*、`代码`。
17. 不要输出空编号、空项目或占位符。如果不知道作品中的角色、规则或运行现象，应直接请学生补充。
18. 不要输出只有一个词的小标题，例如“隐藏”“滑行”“克隆”。需要小标题时，应使用能够表达完整意思的短语。
19. 普通问题原则上控制在3—5句话以内，单次回答一般不超过150字。
20. 截图调试类问题一般控制在2—3句话以内，原则上不超过100字。
21. 不要一次提供多个解决方案。一个方法能够解决当前问题时，不再补充第二种、第三种方法。
22. 只有学生继续追问，或者第一个修改方法没有解决问题时，才进入下一步排查。
23. 不主动扩展新的知识点、复杂技巧或进阶做法，优先帮助学生解决当前问题。
24. 对方向、坐标、广播、变量等概念，只解释解决当前问题所必需的内容，一般用1句话说明。
25. 知识库内容只作为后台参考。回答时要结合当前主题、角色名称和学生表达进行自然改写，不照抄“核心说明、可能原因、学生端引导提示、回答边界”等字段内容。
26. 学生端截图调试中，如果截图已经能够明确定位错误，回答应包含：简短原因、具体修改方法、与当前错误直接相关的局部程序框、运行验证提示。
27. 局部程序框只展示需要修改的这一段角色脚本，不展示其他角色程序，不补充关卡、计时、生命值等无关功能。程序框必须使用以下标记，且标记单独占一行：
[[PROGRAM]]
程序内容
[[/PROGRAM]]

二、关于“怎么做某个作品”的回答规则
1. 当学生问“怎么做某个作品”“这个游戏怎么做”“帮我做某个作品”时，默认判断为任务分析类问题。
2. 第一次回答只能帮助学生理解任务和拆分作品，不得直接给出具体积木名称、积木连接顺序或完整操作步骤。
3. 第一次回答应围绕“作品 → 角色 → 动作、规则、效果”展开。
4. 如果学生只说了作品名称，没有说明角色和功能，应自然地追问1—2个问题，例如作品里有哪些角色、角色需要完成什么动作。
5. 第一次回答结束时，应引导学生先明确或完成一个最基础的小功能，不要同时安排多个复杂功能。
6. 只有学生已经明确目标功能，或者说明自己用了哪些积木、出现了什么问题时，才可以提示关键积木类别。
7. 即使提示积木，也不能一次给出完整程序顺序，只能给一个关键提示，必要时再补充第二个检查方向。
8. 不得使用“第一步拖出……第二步拖出……第三步添加……”这种可以直接照做的完整操作式回答。
9. 对基础任务，应优先引导学生思考：
   ① 动作什么时候开始；
   ② 动作是否需要重复；
   ③ 什么条件会触发变化；
   ④ 运行后应该看到什么效果。

三、关于课堂表现评价
1. 当学生要求“评价我这节课的表现”“评价我的课堂表现”“我这节课表现怎么样”时，不得直接判断学生表现好坏，因为你无法完整观察学生的课堂行为。
2. 应引导学生从任务理解、任务分析、程序搭建、调试修改、合作交流和作品优化等方面进行自我回顾。
3. 可以用简短分点呈现评价维度，但不得直接给学生打分。
4. 不得替代教师进行正式评价。
5. 如果学生说明了自己完成的功能、遇到的问题和解决过程，可以帮助学生整理一段学习表现小结，并提出1—2条具体改进建议。
6. 不要在缺乏依据时直接说“你表现很好”“你完成得很棒”等判断性内容。

四、关于图形化编程功能解释
1. 必须区分“图形化编程内置积木可以直接实现”和“需要组合其他积木或编写额外程序实现”的情况。
2. 不得把不同效果混为一谈。
3. 不主动编造多个复杂实现方案。
4. 如果确实需要说明多个方案，应先介绍最基础、最稳定的方案。只有学生继续追问时，再说明进阶方案。
5. 面向初学者时，不要一开始就引入过难的坐标判断、广播、克隆或复杂变量结构。
6. 使用积木名称时，应尽量采用图形化编程软件中真实显示的名称，避免学生找不到对应积木。

五、关于边缘处理的解释
1. “碰到边缘就反弹”是图形化编程中的内置积木，可以直接实现基础反弹效果。
2. “从舞台一端出去，再从另一端出现”不是一个现成积木，通常需要使用坐标和条件判断实现。
3. 角色碰到边缘后发生倒立或翻转时，应优先检查旋转方式。需要角色保持直立时，通常使用“左右翻转”。
4. 不要把“碰到边缘就反弹”和“从另一端出现”说成同一个功能。
5. 对基础课堂任务，不主动推荐“穿屏出现”等进阶效果。
六、关于具体课堂主题
1. 如果学生问“某个主题怎么做”，应优先引导其完成基础版作品。
2. 不得一开始就把任务扩展成复杂游戏，也不得主动添加教师没有要求的倒计时、关卡、生命值、胜负界面等规则。
3. 当前主题只用于帮助理解学生问题。不同主题中出现的相同程序问题，应使用通用图形化编程知识进行解释。
4. 回答时可以将知识库中的“角色”自然替换成学生当前作品中的具体角色，例如小猫、老鼠、地鼠、锤子或苹果。
七、关于任务分析和逻辑梳理
1. 当学生说“帮我梳理逻辑”“我不知道怎么分析作品”“帮我整理思路”时，应按照“主题 → 背景 → 角色”的结构帮助学生整理。
2. 角色部分是任务分析的重点，每个角色继续分析：
   ① 动作：这个角色需要做什么；
   ② 规则：动作什么时候发生，与哪个角色或条件有关；
   ③ 效果：程序运行后能够看到什么变化。
3. 如果作品没有背景，不要为了结构完整而增加“背景”部分。
4. 不要编造学生没有提到的角色、背景、规则或功能。
5. 如果学生只说了作品主题，应先询问作品中有哪些角色、每个角色需要做什么。
6. 如果学生已经说明了角色和主要功能，可以帮助学生整理成文字版任务分析框架。
7. 任务分析可以采用以下形式，但应根据学生实际内容灵活调整，不要机械照抄：

主题：猫捉老鼠

背景：
房间背景

角色：

小猫

* 动作：根据方向键移动
* 规则：碰到老鼠时触发反馈
* 效果：老鼠被抓到或出现提示

老鼠

* 动作：在舞台中移动
* 规则：被小猫碰到后发生变化
* 效果：隐藏、重新出现或显示结果

八、关于截图调试问题

1. 当学生同时输入文字问题并上传截图时，必须以学生的文字问题为主，结合截图中与该问题直接相关的程序证据回答，不得脱离学生问题另行分析。
2. 当学生只上传截图、没有输入文字问题时，默认任务是“分析这张程序截图中存在哪些问题”。应主动检查截图中能够直接确认的程序问题，不得只做文字识别。
3. 内部分析时必须按以下顺序检查：
   ① 截图中共有几段脚本，每段脚本从哪个事件开始；
   ② 每段脚本中能够明确看见哪些积木、数值和消息名称；
   ③ 学生描述的操作方式是否与条件积木一致，例如“按下鼠标”不能被“碰到鼠标指针”替代；
   ④ 条件成立后执行的动作是否就是学生期望的效果，例如学生希望“地鼠哭”，条件内必须出现对应的造型切换；
   ⑤ 条件是否会持续检测，积木是否放在正确槽位；
   ⑥ 最后才检查多段脚本是否同时修改造型、位置、方向、大小、显示状态或变量，以及是否发生覆盖。
4. 不得根据常见错误自行补全截图。截图中已经存在“等待1秒”时，不得回答“显示和隐藏之间没有等待”；截图中已经存在某个积木时，不得回答“缺少该积木”。
5. 必须区分以下积木：
   ① “等待几秒”的输入框需要数值，它不是等待条件成立；
   ② “等待直到”用于等待条件成立；
   ③ “如果……那么……”用于条件成立后执行一组积木；
   ④ “循环执行”会持续运行，放在点击事件下也不会自动结束。
6. 如果同一角色存在多段脚本，应检查它们是否并发修改同一属性，但多脚本冲突只能放在诊断后段。必须先检查：学生期望的触发方式是否与条件积木一致、条件内是否包含学生期望的反馈动作、条件是否会持续检测；只有这些都正确时，才把脚本冲突作为主要原因。
7. 当学生提出了明确问题时，只指出与该问题关系最直接、证据最充分的一个主要问题，并给出一个最相关的修改方向。
8. 当学生只上传截图时，先指出截图中最主要、最确定的问题；如果截图还明确显示第二个与程序运行直接相关的问题，可以再简要指出第二个问题。不要一次罗列大量可能原因。
9. 每个确认的问题只给一个修改方向。修改方向必须具体说明“现有程序中的哪个积木要怎样改”，不能只说“删掉脚本”“只保留另一段脚本”“重新写一遍”。应优先保留学生已有程序，例如提示“把循环执行改成一次执行”“把等待中的条件改为固定秒数”“把条件放入如果……那么”。
9.1 回答和程序框必须逐字保留学生截图中每一个已有积木的原文，不只包括“重复执行/循环执行”，还包括事件、条件、运算符、角色名、碰撞对象、造型名、消息名、变量名、数值和动作积木。不得自行统一术语或用近义表达替换。
9.2 修改程序时，应优先复制学生原脚本，只调整积木位置、嵌套层级或当前错误对应的参数。只有确实缺少解决当前问题所必需的积木时，才允许新增；新增积木优先采用同一截图中已经出现的平台术语。
9.3 如果视觉证据不足以确认某个积木的准确名称、参数或连接关系，不得生成程序框，应只追问一个关键问题。
10. 当截图已经能够明确定位错误时，应在原因和修改说明后，展示一段“修改后的关键程序”。只展示与当前问题直接相关的脚本，不展示整个作品。程序框使用 [[PROGRAM]] 与 [[/PROGRAM]] 包住，并保持积木的缩进层级。
11. 只有当整段脚本与任务完全无关、且无法通过调整其中积木修复时，才可以建议删除；一般调试问题不得把“删除整段脚本”作为首选建议。
12. 如果当前问题有多个明确关联的错误，可以在同一个局部程序框中给出能够直接修复当前现象的正确结构；仍不得一次给出所有角色的完整程序。
13. 如果截图结构分析与OCR文字相互矛盾，或者积木被遮挡、文字过小、连接关系看不清，不要猜测。应说明当前截图能够确认的内容，并只追问一个关键问题。
14. 如果学生只上传截图，但无法判断其作品目标或预期效果，可以先指出截图中能够直接确认的结构问题；如果没有可确认的问题，再追问“你希望程序实现什么效果？”或“运行后出现了什么现象？”。
15. 当截图中已经能够判断错误设置和正确修改值时，应给出唯一、明确的修改建议，不再补充无关备选方案。
16. 截图调试回答按“原因说明—正确做法—局部程序框—运行验证”组织。程序框外的文字应简短，整段回答以学生能够快速看懂为准。
17. 调试类回答应像教师在学生身边进行即时指导，不要写成错误分析报告，也不要在结尾总结“核心问题是……”。

九、自然对话表达规则
1. 回答应像教师在学生旁边进行简短指导，不要像说明书、知识点讲解或评价报告。
1.1 回答对象就是当前学生本人，必须使用第二人称表达，例如“你把点击触发的动作放在了……”“把你的这段脚本改成……”。禁止写成“学生把……”“学生需要……”或“该学生……”。
2. 学生问题已经明确时，直接回答，不重复学生的原话，不使用“你的问题是……”“人歪了”“角色不动了”等生硬开头。
3. 简单问题优先使用1—2句话回答。第一句话指出截图中能够确认的设置或现象，第二句话说明最关键的修改方法。
4. 不强制使用分段、编号或固定结构。只有确实存在多个步骤时，才使用①②③。
5. 不主动讲解大段积木原理；但如果回答中直接给出坐标、角度、秒数、次数、变量值等具体数值，必须顺手补一句这个数值为什么这样填。
6. 回答时应使用截图中具体的信息，例如：
   “角色方向现在是60°”
   “程序中使用的是面向鼠标指针”
   “显示和隐藏之间没有等待”
   不要使用“方向积木控制角色朝哪个角度”等抽象、书面化表达。
7. 不要照搬学生的口语表达。例如学生说“人为什么是歪的”，回答时应使用“角色方向”“人物显示”等准确表达，不重复说“人歪了”。
8. 不要每次使用“你可以这样调整”“试试看”“再运行看看”“就可以了”等相同句式。根据问题自然结束回答。
9. 截图能够明确判断时，直接给出结论和修改方法，不使用“可能是”“通常是”等模糊表达。
10. 截图不能明确判断时，应使用“从目前截图还不能确定”“先检查……”等谨慎表达，并只追问一个关键问题。
11. 知识库只提供判断依据。回答时必须重新组织语言，不得照抄知识库中的“学生端引导提示”。
12. 截图结构分析只作为后台证据，不要把“各段脚本、视觉分析结果”等内部分析全文展示给学生。
13. 学生端可以使用程序框展示“当前错误对应的修改后局部脚本”。程序框应像积木从上到下的连接顺序一样缩进，便于学生对照自己的程序修改；不得用程序框展示整个作品。
14. 学生问“填多少”“是多少”“应该填几”时，回答不能只给数字；要先给数字，再用截图里的起点、终点、增量或舞台位置解释一句。
15. 回答中不添加无关的问候、鼓励、总结或扩展知识。解决当前问题后即可结束。

"""


# =========================
# 3. 日志保存
# =========================

def save_log(
    role: str,
    user_input: str,
    answer: str,
    student_name="",
    group_no="",
    topic="",
    uploaded_image_name="",
    uploaded_image_path="",
    ocr_text="",
    screenshot_analysis="",
    uploaded_image_base64=""
):
    os.makedirs("logs", exist_ok=True)
    log_path = "logs/chat_logs.csv"

    new_row = {
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "role": role,
        "student_name": student_name,
        "group_no": group_no,
        "topic": topic,
        "user_input": user_input,
        "answer": answer,
        "uploaded_image_name": uploaded_image_name,
        "uploaded_image_path": uploaded_image_path,
        "ocr_text": ocr_text,
        "screenshot_analysis": screenshot_analysis,
        "uploaded_image_base64": uploaded_image_base64
    }

    if os.path.exists(log_path):
        df = pd.read_csv(log_path)
        # 兼容旧版本日志：如果旧 CSV 没有新增列，自动补上空列。
        for column in new_row.keys():
            if column not in df.columns:
                df[column] = ""
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    else:
        df = pd.DataFrame([new_row])

    df.to_csv(log_path, index=False, encoding="utf-8-sig")


def load_logs():
    log_path = "logs/chat_logs.csv"
    columns = [
        "time", "role", "student_name", "group_no", "topic",
        "user_input", "answer", "uploaded_image_name", "uploaded_image_path", "ocr_text", "screenshot_analysis", "uploaded_image_base64"
    ]

    if not os.path.exists(log_path):
        return pd.DataFrame(columns=columns)

    if os.path.getsize(log_path) == 0:
        return pd.DataFrame(columns=columns)

    try:
        df = pd.read_csv(log_path, encoding="utf-8-sig", on_bad_lines="skip")
    except Exception:
        try:
            df = pd.read_csv(log_path, encoding="utf-8-sig", engine="python", on_bad_lines="skip")
        except Exception:
            return pd.DataFrame(columns=columns)

    for column in columns:
        if column not in df.columns:
            df[column] = ""

    return df


def save_uploaded_image_file(uploaded_file):
    """
    将学生上传的图片保存到 logs/uploads 文件夹。
    CSV 中会记录图片文件名和相对路径；教师下载 ZIP 时会一起下载原图。
    """
    if uploaded_file is None:
        return "", ""

    upload_dir = os.path.join("logs", "uploads")
    os.makedirs(upload_dir, exist_ok=True)

    original_name = os.path.basename(uploaded_file.name or "uploaded_image.png")
    safe_name = re.sub(r"[^0-9a-zA-Z_\-.\u4e00-\u9fff]", "_", original_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    saved_name = f"{timestamp}_{safe_name}"
    saved_path = os.path.join(upload_dir, saved_name)

    uploaded_file.seek(0)
    with open(saved_path, "wb") as f:
        f.write(uploaded_file.getvalue())
    uploaded_file.seek(0)

    return saved_name, saved_path


def build_logs_zip_bytes():
    """
    打包下载全部对话记录和上传图片。
    ZIP 中包含：
    1. chat_logs.csv
    2. uploads/ 文件夹中的原始截图
    """
    buffer = io.BytesIO()
    log_path = os.path.join("logs", "chat_logs.csv")
    upload_dir = os.path.join("logs", "uploads")

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        if os.path.exists(log_path):
            zip_file.write(log_path, arcname="chat_logs.csv")

        if os.path.exists(upload_dir):
            for root, _, files in os.walk(upload_dir):
                for filename in files:
                    file_path = os.path.join(root, filename)
                    arcname = os.path.relpath(file_path, "logs")
                    zip_file.write(file_path, arcname=arcname)

    buffer.seek(0)
    return buffer.getvalue()


def create_excel_with_images(logs_df):
    """
    生成包含全部对话记录的 Excel。
    每一行对应一次用户提问和智能体回答；
    上传图片会嵌入到“上传截图”列，教师打开 Excel 可以直接看到图片。
    新记录会同时保存图片路径和 base64；如果路径失效，会尝试从 base64 恢复图片。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "对话记录"

    headers = [
        "时间", "身份", "小组号", "当前主题",
        "学生/教师问题", "上传截图", "智能体回答", "OCR识别文字", "截图结构分析"
    ]
    ws.append(headers)

    column_widths = {
        "A": 20,
        "B": 12,
        "C": 12,
        "D": 16,
        "E": 42,
        "F": 36,
        "G": 70,
        "H": 48,
        "I": 58,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    def safe_cell(value):
        if value is None:
            return ""
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
        return str(value)

    def clean_question_for_excel(value):
        """Excel 的问题列中不再重复显示“上传图片：xxx”，图片由截图列直接展示。"""
        text = safe_cell(value)
        text = re.sub(r"\n*（已上传图片：.*?）", "", text)
        text = re.sub(r"（上传图片：.*?）", "", text)
        return text.strip()

    def resolve_image_path(raw_path):
        """兼容相对路径、旧路径、Windows/Unix 分隔符。"""
        raw_path = safe_cell(raw_path).strip()
        if not raw_path or raw_path.lower() == "nan":
            return ""

        normalized = raw_path.replace("\\", os.sep).replace("/", os.sep)
        candidates = [
            raw_path,
            normalized,
            os.path.abspath(raw_path),
            os.path.abspath(normalized),
            os.path.join(os.getcwd(), raw_path),
            os.path.join(os.getcwd(), normalized),
        ]

        # 如果 CSV 里只保存了文件名，也尝试到 logs/uploads 里找
        basename = os.path.basename(normalized)
        if basename:
            candidates.append(os.path.join("logs", "uploads", basename))
            candidates.append(os.path.join(os.getcwd(), "logs", "uploads", basename))

        seen = set()
        for candidate in candidates:
            if not candidate or candidate in seen:
                continue
            seen.add(candidate)
            if os.path.exists(candidate):
                return candidate
        return ""

    def normalize_image_to_temp_file(image_path="", image_b64="", row_no=0):
        """
        将图片统一转成临时 PNG 文件，再交给 openpyxl 插入。
        这样比直接把 BytesIO 交给 ExcelImage 更稳定。
        """
        try:
            if image_path and os.path.exists(image_path):
                pil_img = PILImage.open(image_path)
            elif image_b64:
                raw = base64.b64decode(image_b64)
                pil_img = PILImage.open(io.BytesIO(raw))
            else:
                return ""

            pil_img = pil_img.convert("RGB")
            pil_img.thumbnail((260, 180))

            temp_dir = os.path.join("logs", "excel_temp_images")
            os.makedirs(temp_dir, exist_ok=True)
            temp_path = os.path.join(temp_dir, f"excel_row_{row_no}.png")
            pil_img.save(temp_path, format="PNG")
            return temp_path
        except Exception:
            return ""

    for idx, row in logs_df.iterrows():
        excel_row = idx + 2

        ws.cell(excel_row, 1, safe_cell(row.get("time", "")))
        ws.cell(excel_row, 2, safe_cell(row.get("role", "")))
        ws.cell(excel_row, 3, safe_cell(row.get("group_no", "")))
        ws.cell(excel_row, 4, safe_cell(row.get("topic", "")))
        ws.cell(excel_row, 5, clean_question_for_excel(row.get("user_input", "")))
        ws.cell(excel_row, 7, safe_cell(row.get("answer", "")))
        ws.cell(excel_row, 8, safe_cell(row.get("ocr_text", "")))
        ws.cell(excel_row, 9, safe_cell(row.get("screenshot_analysis", "")))

        # 行高要足够，否则图片虽然插入了，但看起来像“没显示”。
        ws.row_dimensions[excel_row].height = 145

        image_path = resolve_image_path(row.get("uploaded_image_path", ""))
        image_b64 = safe_cell(row.get("uploaded_image_base64", ""))
        image_name = safe_cell(row.get("uploaded_image_name", ""))

        temp_image_path = normalize_image_to_temp_file(image_path, image_b64, excel_row)
        if temp_image_path:
            try:
                excel_img = ExcelImage(temp_image_path)
                excel_img.width = 240
                excel_img.height = 160
                ws.add_image(excel_img, f"F{excel_row}")
            except Exception:
                ws.cell(excel_row, 6, f"图片插入失败：{image_name or image_path}")
        elif image_name:
            ws.cell(excel_row, 6, f"图片文件未找到：{image_name}")
        else:
            ws.cell(excel_row, 6, "")

        for col_idx in range(1, 10):
            ws.cell(excel_row, col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =========================
# 4. 清理模型回复
# =========================

def clean_answer(text: str) -> str:
    """
    清理模型回复：
    1. 删除不希望出现的 HTML 标签；
    2. 删除流式输出残留符号；
    3. 删除空编号、空项目；
    4. 清理多余空行。
    注意：这里不再删除 Markdown 表格竖线，也不强制把所有编号转成①②③。
    学生端编号统一交给 normalize_student_numbering()；
    教师端保留（一）/1./①等层级编号，并允许教学过程表格正常显示。
    """
    if not text:
        return ""

    for bad_char in ["▌", "█", "▐", "▍", "▎", "▏", "■", "□", "▪", "▫"]:
        text = text.replace(bad_char, "")

    text = text.replace("**", "")
    text = text.replace("*", "")
    text = text.replace("`", "")

    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</p\s*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<p\s*>", "", text, flags=re.IGNORECASE)
    text = re.sub(r"</div\s*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<div\s*>", "", text, flags=re.IGNORECASE)
    text = re.sub(r"</?span[^>]*>", "", text, flags=re.IGNORECASE)
    text = re.sub(r"</?strong[^>]*>", "", text, flags=re.IGNORECASE)
    text = re.sub(r"</?b[^>]*>", "", text, flags=re.IGNORECASE)
    text = text.replace("&nbsp;", " ")

    text = re.sub(r"^\s*[①②③④⑤⑥⑦⑧⑨⑩]\s*$", "", text, flags=re.MULTILINE)
    text = re.sub(r"^\s*[（(]?\d{1,2}[）).、．]\s*$", "", text, flags=re.MULTILINE)

    incomplete_words = r"被|隐藏|显示|滑行|克隆|碰到|点击|分数|变量|造型|位置|条件"
    text = re.sub(rf'^\s*[①②③④⑤⑥⑦⑧⑨⑩]\s*[‘’\'"“”]?({incomplete_words})[‘’\'"“”]?\s*$', "", text, flags=re.MULTILINE)
    text = re.sub(rf'^\s*[（(]?\d{{1,2}}[）).、．]\s*[‘’\'"“”]?({incomplete_words})[‘’\'"“”]?\s*$', "", text, flags=re.MULTILINE)

    text = re.sub(r"^\s*[:：]\s*$", "", text, flags=re.MULTILINE)

    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)

    return text.strip()

def normalize_student_numbering(text: str) -> str:
    """
    只用于学生端：把常见的 1.、（1）、(1)、1、 等编号统一成 ①②③。
    教师端不要调用这个函数，避免破坏正式层级编号。
    """
    if not text:
        return ""

    number_map = [
        ("10", "⑩"),
        ("1", "①"),
        ("2", "②"),
        ("3", "③"),
        ("4", "④"),
        ("5", "⑤"),
        ("6", "⑥"),
        ("7", "⑦"),
        ("8", "⑧"),
        ("9", "⑨"),
    ]

    for num, circled in number_map:
        text = text.replace(f"（{num}）", circled)
        text = text.replace(f"({num})", circled)

    converted_lines = []

    for line in text.splitlines():
        stripped = line.lstrip()
        prefix = line[:len(line) - len(stripped)]
        new_line = line

        for num, circled in number_map:
            for mark in [".", "．", "、", ")"]:
                marker = f"{num}{mark}"
                if stripped.startswith(marker):
                    new_line = prefix + circled + " " + stripped[len(marker):].lstrip()
                    break

            if new_line != line:
                break

        converted_lines.append(new_line)

    text = "\n".join(converted_lines)
    text = re.sub(r"\n{3,}", "\n\n", text)

    return text.strip()


def postprocess_answer(role: str, text: str) -> str:
    """
    统一清理模型回答。
    学生端额外统一编号；教师端保留层级编号。
    """
    answer = clean_answer(text)

    if role == "学生端":
        answer = normalize_student_numbering(answer)
        # 学生端是与当前学生直接对话，避免模型偶尔使用旁观式的第三人称表述。
        direct_address_replacements = {
            "该学生的程序": "你的程序",
            "这名学生的程序": "你的程序",
            "学生的程序": "你的程序",
            "该学生把": "你把",
            "这名学生把": "你把",
            "学生把": "你把",
            "该学生将": "你将",
            "这名学生将": "你将",
            "学生将": "你将",
            "该学生需要": "你需要",
            "这名学生需要": "你需要",
            "学生需要": "你需要",
        }
        for source, replacement in direct_address_replacements.items():
            answer = answer.replace(source, replacement)

    return answer

def apply_simple_bold(text: str) -> str:
    parts = text.split("**")
    if len(parts) < 3:
        return text

    result = []
    for index, part in enumerate(parts):
        if index % 2 == 1:
            result.append("<strong>" + part + "</strong>")
        else:
            result.append(part)

    return "".join(result)


def format_message_content(content: str) -> str:
    """
    将模型文本转换成更适合气泡显示的 HTML。
    支持普通分行、标题、项目符号、编号行，以及教师端教学过程常用的 Markdown 表格。
    """
    content = content or ""

    circled_numbers = ["①", "②", "③", "④", "⑤", "⑥", "⑦", "⑧", "⑨", "⑩"]
    parts = []
    lines = content.splitlines()
    i = 0

    def is_table_line(line: str) -> bool:
        s = line.strip()
        return s.startswith("|") and s.endswith("|") and s.count("|") >= 2

    def is_separator_line(line: str) -> bool:
        s = line.strip().strip("|").strip()
        if not s:
            return False
        cells = [c.strip() for c in s.split("|")]
        return all(re.fullmatch(r":?-{3,}:?", c or "") for c in cells)

    def split_table_cells(line: str):
        return [html.escape(c.strip()) for c in line.strip().strip("|").split("|")]

    def render_table(table_lines):
        if len(table_lines) < 2:
            return ""
        header = split_table_cells(table_lines[0])
        body_start = 1
        if len(table_lines) > 1 and is_separator_line(table_lines[1]):
            body_start = 2
        rows = [split_table_cells(row) for row in table_lines[body_start:] if is_table_line(row)]
        html_parts = ['<div class="md-table-wrap"><table class="md-table">']
        html_parts.append("<thead><tr>")
        for cell in header:
            html_parts.append(f"<th>{cell}</th>")
        html_parts.append("</tr></thead>")
        if rows:
            html_parts.append("<tbody>")
            for row in rows:
                html_parts.append("<tr>")
                for idx in range(len(header)):
                    cell = row[idx] if idx < len(row) else ""
                    html_parts.append(f"<td>{cell}</td>")
                html_parts.append("</tr>")
            html_parts.append("</tbody>")
        html_parts.append("</table></div>")
        return "".join(html_parts)

    while i < len(lines):
        line = lines[i]
        stripped_raw = line.strip()

        if stripped_raw == "[[PROGRAM]]":
            program_lines = []
            i += 1
            while i < len(lines) and lines[i].strip() != "[[/PROGRAM]]":
                program_lines.append(lines[i].rstrip())
                i += 1
            if i < len(lines) and lines[i].strip() == "[[/PROGRAM]]":
                i += 1
            program_text = html.escape("\n".join(program_lines).strip("\n"))
            parts.append('<div class="program-block"><pre>' + program_text + '</pre></div>')
            continue

        if is_table_line(line):
            table_lines = []
            while i < len(lines) and (is_table_line(lines[i]) or is_separator_line(lines[i])):
                table_lines.append(lines[i])
                i += 1
            table_html = render_table(table_lines)
            if table_html:
                parts.append(table_html)
            continue

        stripped = html.escape(stripped_raw)
        stripped = apply_simple_bold(stripped)

        if not stripped_raw:
            parts.append('<div class="md-space"></div>')
            i += 1
            continue

        if stripped_raw.startswith("#"):
            heading_text = html.escape(stripped_raw.lstrip("#").strip())
            if heading_text:
                parts.append('<div class="md-heading">' + heading_text + '</div>')
                i += 1
                continue

        if stripped_raw.startswith("- "):
            parts.append('<div class="md-bullet">• ' + html.escape(stripped_raw[2:].strip()) + '</div>')
            i += 1
            continue

        if stripped_raw.startswith("• "):
            parts.append('<div class="md-bullet">' + html.escape(stripped_raw) + '</div>')
            i += 1
            continue

        is_number_line = False
        for number_text in circled_numbers:
            if stripped_raw.startswith(number_text):
                is_number_line = True
                break

        for number in range(1, 21):
            if stripped_raw.startswith(str(number) + ".") or stripped_raw.startswith(str(number) + "．") or stripped_raw.startswith(str(number) + "、"):
                is_number_line = True
                break

        if re.match(r"^[（(]\d{1,2}[）)]", stripped_raw):
            is_number_line = True

        if is_number_line:
            parts.append('<div class="md-number">' + stripped + '</div>')
        else:
            parts.append('<div class="md-line">' + stripped + '</div>')

        i += 1

    return "".join(parts)


def build_chat_bubble_html(role: str, content: str, image_base64: str = "") -> str:
    """
    构建聊天气泡 HTML。
    如果用户消息带有截图，则把截图嵌入到同一个气泡中，避免“上一张图片消失”。
    助手消息在最终渲染前再统一一次编号，避免出现 ①②③（4）⑤ 混用。
    """
    display_content = content or ""
    safe_content = format_message_content(display_content)

    image_html = ""
    if image_base64:
        image_html = f"""
        <img class="chat-image" src="data:image/png;base64,{image_base64}" />
        """

    if role == "user":
        bubble_role = "user"
        avatar = "我"
    else:
        bubble_role = "assistant"
        avatar = "🐱"

    return f"""
    <div class="chat-wrap {bubble_role}">
        <div class="chat-avatar {bubble_role}">{avatar}</div>
        <div class="chat-content">
            <div class="chat-bubble {bubble_role}">{safe_content}{image_html}</div>
        </div>
    </div>
    """


def render_chat_bubble(role: str, content: str, image_base64: str = ""):
    """
    使用自定义 HTML 渲染类似微信的左右聊天气泡。
    """
    st.markdown(
        build_chat_bubble_html(role, content, image_base64),
        unsafe_allow_html=True
    )

def scroll_to_bottom(holder=None, smooth: bool = True):
    """
    尽量让页面自动滚动到底部。
    这个版本只滚动页面和主容器，不再强制 chat_input 进入视野，
    避免右侧滚动条反复向上/向下跳动。
    """
    behavior = "smooth" if smooth else "auto"
    stamp = str(time.time()).replace(".", "")

    script = f"""
    <script id="scroll-bottom-{stamp}">
    function scrollBottom{stamp}() {{
        try {{
            const doc = window.parent.document;
            const targets = [];

            if (doc.scrollingElement) targets.push(doc.scrollingElement);
            if (doc.documentElement) targets.push(doc.documentElement);
            if (doc.body) targets.push(doc.body);

            const selectors = [
                '[data-testid="stAppViewContainer"]',
                '[data-testid="stMain"]',
                'section.main',
                '.main'
            ];

            selectors.forEach(function(sel) {{
                const el = doc.querySelector(sel);
                if (el) targets.push(el);
            }});

            targets.forEach(function(el) {{
                try {{
                    const maxTop = Math.max(el.scrollHeight, el.offsetHeight, el.clientHeight) + 2000;
                    if (typeof el.scrollTo === "function") {{
                        el.scrollTo({{ top: maxTop, behavior: "{behavior}" }});
                    }}
                    el.scrollTop = maxTop;
                }} catch (e) {{}}
            }});
        }} catch (e) {{}}
    }}

    setTimeout(scrollBottom{stamp}, 60);
    setTimeout(scrollBottom{stamp}, 180);
    setTimeout(scrollBottom{stamp}, 420);
    setTimeout(scrollBottom{stamp}, 800);
    </script>
    """

    if holder is not None:
        with holder.container():
            components.html(script, height=0)
    else:
        components.html(script, height=0)


# =========================
# 4.5 图片文字识别 OCR
# =========================
def analyze_sb3_file(sb3_file):
    """
    解析图形化编程.sb3 文件，提取用于作品评价和与教师基础作品对比的结构信息。
    """
    result = {
        "file_name": sb3_file.name,
        "sprites": [],
        "stage_count": 0,
        "sprite_count": 0,
        "block_count": 0,
        "event_blocks": 0,
        "loop_blocks": 0,
        "condition_blocks": 0,
        "variable_blocks": 0,
        "broadcast_blocks": 0,
        "looks_blocks": 0,
        "motion_blocks": 0,
        "sound_blocks": 0,
        "operators_blocks": 0,
        "sensing_blocks": 0,
        "opcodes": [],
        "summary": ""
    }

    try:
        sb3_file.seek(0)
        with zipfile.ZipFile(sb3_file, "r") as z:
            with z.open("project.json") as f:
                project = json.load(f)

        targets = project.get("targets", [])

        for target in targets:
            name = target.get("name", "")
            is_stage = target.get("isStage", False)
            blocks = target.get("blocks", {})

            if is_stage:
                result["stage_count"] += 1
            else:
                result["sprite_count"] += 1
                result["sprites"].append(name)

            for block_id, block in blocks.items():
                if not isinstance(block, dict):
                    continue

                opcode = block.get("opcode", "")
                if not opcode:
                    continue

                result["block_count"] += 1
                result["opcodes"].append(opcode)

                if opcode.startswith("event_"):
                    result["event_blocks"] += 1
                elif opcode.startswith("control_repeat") or opcode.startswith("control_forever"):
                    result["loop_blocks"] += 1
                elif opcode.startswith("control_if"):
                    result["condition_blocks"] += 1
                elif opcode.startswith("data_"):
                    result["variable_blocks"] += 1
                elif "broadcast" in opcode:
                    result["broadcast_blocks"] += 1
                elif opcode.startswith("looks_"):
                    result["looks_blocks"] += 1
                elif opcode.startswith("motion_"):
                    result["motion_blocks"] += 1
                elif opcode.startswith("sound_"):
                    result["sound_blocks"] += 1
                elif opcode.startswith("operator_"):
                    result["operators_blocks"] += 1
                elif opcode.startswith("sensing_"):
                    result["sensing_blocks"] += 1

        result["opcodes"] = sorted(list(set(result["opcodes"])))

        result["summary"] = (
            f"作品包含 {result['sprite_count']} 个角色，"
            f"共检测到 {result['block_count']} 个积木；"
            f"事件积木 {result['event_blocks']} 个，"
            f"循环积木 {result['loop_blocks']} 个，"
            f"条件积木 {result['condition_blocks']} 个，"
            f"变量相关积木 {result['variable_blocks']} 个，"
            f"广播相关积木 {result['broadcast_blocks']} 个。"
        )

    except Exception as e:
        result["summary"] = f"解析失败：{e}"

    return result


def extract_eval_field(text, field_name):
    """
    从模型输出中提取指定字段。
    支持字段前带编号，也支持多行字段内容。
    """
    if not text:
        return ""

    field_names = [
        "核心功能完成情况",
        "完成等级",
        "各维度评分及依据",
        "建议总分",
        "运行画面分析",
        "总评依据",
        "评分依据",
        "改进建议",
        "教师复核点",
        # 兼容旧版本输出
        "维度评分",
        "完整性建议分",
        "技术性建议分",
        "创新性建议分",
        "艺术性建议分",
    ]

    field_pattern = r"(?:[（(]?[一二三四五六七八九十\d]+[）)]?[、.．]?\s*)?"
    all_fields = "|".join(map(re.escape, field_names))

    pattern = (
        field_pattern
        + re.escape(field_name)
        + r"\s*[:：]\s*"
        + r"(.*?)"
        + r"(?=\n\s*"
        + field_pattern
        + r"(?:"
        + all_fields
        + r")\s*[:：]|\Z)"
    )

    match = re.search(pattern, text, flags=re.S)

    if not match:
        return ""

    value = match.group(1).strip()
    value = re.sub(r"\n{2,}", "\n", value)
    return value


def extract_score_number(text):
    """
    从“25分”“建议25”“25/30”等文本中提取第一个数字。
    用于写入 Excel 的总分字段。
    """
    if not text:
        return ""

    match = re.search(r"\d+(?:\.\d+)?", str(text))
    if not match:
        return ""

    number = match.group(0)

    if "." in number:
        return float(number)

    return int(number)


def normalize_completion_level(level_text):
    """
    完成等级只保留等级名称，不保留括号、分数区间或解释。
    """
    text = str(level_text or "").strip()
    for level in ["优秀完成", "良好完成", "基本完成", "部分完成", "未完成"]:
        if level in text:
            return level
    return text


def parse_dimension_evaluations(eval_text):
    """
    解析“各维度评分及依据”字段。
    期望模型输出格式：
    各维度评分及依据：
    完整性评分及依据：26/30。依据：……
    技术性评分及依据：25/30。依据：……

    返回：
    {
        "完整性评分及依据": "26/30。依据：……",
        "技术性评分及依据": "25/30。依据：……"
    }
    """
    result = {}

    block = extract_eval_field(eval_text, "各维度评分及依据")
    if not block:
        # 兼容旧版本“维度评分”字段
        block = extract_eval_field(eval_text, "维度评分")

    if not block:
        return result

    lines = [line.strip() for line in block.splitlines() if line.strip()]

    current_key = ""
    current_value_parts = []

    def flush_current():
        if current_key:
            value = " ".join(part.strip() for part in current_value_parts if part.strip()).strip()
            if value:
                result[current_key] = value

    for line in lines:
        # 支持：完整性评分及依据：26/30。依据：……
        m = re.match(r"^(.{1,30}?评分及依据)\s*[:：]\s*(.+)$", line)
        if m:
            flush_current()
            current_key = m.group(1).strip()
            current_value_parts = [m.group(2).strip()]
            continue

        # 支持：完整性：26/30。依据：……
        m = re.match(r"^(.{1,20}?)\s*[:：]\s*(\d+(?:\.\d+)?\s*/\s*\d+.*)$", line)
        if m:
            flush_current()
            dim_name = m.group(1).strip()
            if not dim_name.endswith("评分及依据"):
                dim_name = f"{dim_name}评分及依据"
            current_key = dim_name
            current_value_parts = [m.group(2).strip()]
            continue

        # 支持多行依据接在上一行后面
        if current_key:
            current_value_parts.append(line)

    flush_current()
    return result


TEACHER_CASE_RULES = """
【教师评价案例集形成的评分尺度】

本评分尺度来自研究者与信息科技教师共同评价学生作品后形成的评价案例集。
该规则只用于帮助智能体判断完成等级和评分尺度，不用于决定评价维度。

一、完成等级与总分区间
1. 优秀完成：90—95分
核心功能全部完成，程序运行稳定，没有明显错误，并有一定优化、拓展或创意表现。

2. 良好完成：85—89分
核心功能基本完整，程序能够正常运行，但存在较小缺陷、程序结构问题或表现不够完善。

3. 基本完成：75—84分
完成了大部分核心功能，但缺失一项关键功能，或部分交互、反馈、运行效果不完整。

4. 部分完成：60—74分
只完成了部分核心功能，作品能够呈现一定效果，但尚未形成完整作品逻辑。

5. 未完成：35—59分
核心功能未实现，程序无法形成有效作品，或主要角色、交互、运行逻辑明显缺失。

二、评分原则
1. 评价时必须优先判断本课核心功能完成情况。
2. 教师上传的评价量表决定评价维度和各维度分值。
3. 教师评价案例集只用于确定评分尺度，不得替代教师上传的评价量表。
4. 如果课堂任务或评价量表没有明确要求变量、广播、倒计时、胜负界面、关卡切换，不得因为缺失这些功能直接扣分。
5. 如果学生作品实现了角色控制、移动、碰撞检测、消息反馈、造型变化、重新出现等核心交互，应给予合理肯定。
6. 如果作品存在明显运行错误，例如角色方向错误、无法控制、核心角色不动、碰撞后无反馈、程序只能运行一次，应根据影响程度降低完成等级。
7. 如果视频关键帧没有完整呈现某项功能，不能直接判定该功能缺失，应写入“教师复核点”。
8. 如果学生作品已经实现本课核心功能，程序能够运行，且作品能够表达任务目标，一般应评为良好完成或优秀完成，不应仅因缺少扩展功能而降为基本完成。
9. 如果学生作品虽然能看到主要效果，但存在“只能运行一次、抓到后不能重新开始、角色控制方式明显不符合任务、关键反馈缺失、运行逻辑不稳定”等问题，应降为基本完成或部分完成。
"""


def load_teacher_cases_text():
    """
    读取后端内置的教师评价案例集。
    文件名固定为 teacher_cases.xlsx，放在和本 .py 文件同一目录。
    案例集用于帮助智能体学习教师评分尺度，不要求普通教师每次上传。
    """
    case_path = APP_DIR / "teacher_cases.xlsx"

    if not case_path.exists():
        return "未检测到 teacher_cases.xlsx，本次仅依据评分规则和教师上传评价量表进行评价。"

    try:
        df = pd.read_excel(case_path)
    except Exception as e:
        return f"teacher_cases.xlsx 读取失败：{e}"

    needed_cols = ["作品编号", "主题","完成等级", "建议总分", "评分依据", "改进建议"]
    for col in needed_cols:
        if col not in df.columns:
            df[col] = ""

    case_lines = []
    for _, row in df.iterrows():
        case_text = (
            f"案例{row.get('作品编号', '')}\n"
            f"主题：{row.get('主题', '')}\n"
            f"完成等级：{row.get('完成等级', '')}\n"
            f"教师评分：{row.get('建议总分', '')}\n"
            f"评分依据：{row.get('评分依据', '')}\n"
            f"改进建议：{row.get('改进建议', '')}\n"
        )
        case_lines.append(case_text)

    return "\n---\n".join(case_lines[:40])


def evaluate_project_with_rubric(sb3_analysis, rubric_text, reference_analysis=None, video_analysis=None, reference_video_analysis=None):
    """
    根据评价量表、教师基础作品、教师基础版运行视频、学生作品结构和学生运行视频生成辅助评价。
    教师基础版运行视频用于提供当前任务的“可见运行效果参照”，不再要求教师填写文字版运行效果说明。
    """
    reference_text = ""
    teacher_cases_text = load_teacher_cases_text()

    if reference_analysis or reference_video_analysis:
        reference_text = f"""
【教师基础作品参照】
教师基础作品和教师基础版运行视频只作为“核心功能目标参照”，不是标准答案代码。
学生不需要和教师作品使用完全相同的积木或实现方式，只要实现相同或相近的核心运行效果即可。

【教师基础作品结构】
{json.dumps(reference_analysis, ensure_ascii=False, indent=2) if reference_analysis else "未上传教师基础作品 .sb3。"}

【教师基础版运行视频分析】
以下内容来自 Qwen-VL 对教师基础版作品运行视频关键帧的观察，用于帮助判断本节课作品应呈现的核心运行效果。
{json.dumps(reference_video_analysis, ensure_ascii=False, indent=2) if reference_video_analysis else "未上传教师基础版运行视频。"}
"""

    if video_analysis:
        video_text = f"""
【学生作品运行视频分析】
以下内容来自 Qwen-VL 对学生作品运行录屏关键帧的观察，只作为运行效果参考。
如果视频分析与 .sb3 结构分析不一致，请在“教师复核点”中提示教师进一步核对。
{json.dumps(video_analysis, ensure_ascii=False, indent=2)}
"""
    else:
        video_text = """
【学生作品运行视频分析】
未上传学生作品运行视频，本次评价主要依据学生作品结构分析、教师基础作品、教师基础版运行视频和评价量表。
"""

    prompt = f"""
你是小学五年级图形化编程作品辅助评价助手。

请根据教师上传的作品评价量表、学生作品结构分析结果、学生作品运行视频分析、教师基础作品结构、教师基础版运行视频分析，以及教师评价案例集形成的评分尺度，生成作品辅助评价建议。

【最重要的评价规则】
1. 必须严格依据教师上传的评价量表进行评价。
2. 评价维度必须来自教师上传的评价量表，不得固定使用“完整性、技术性、创新性、艺术性”等系统预设维度。
3. 每个评价维度都必须给出“建议分数”和“评分依据”。
4. 每个维度的评分依据必须对应该维度在教师上传评价量表中的评价要求。
5. 如果教师上传的评价量表包含“任务完成度、编程思维、作品表达”等维度，就必须按这些维度评分。
6. 如果教师上传的评价量表包含“完整性、技术性、创新性、艺术性”等维度，才可以按这些维度评分。
7. 教师评价案例集只用于学习评分尺度和完成等级。
8. 教师基础作品和教师基础版运行视频共同作为核心功能目标参照，不是标准答案代码。
9. 教师基础版运行视频用于帮助理解本节课基础作品应呈现的可见效果；不得把教师基础版作品中的所有细节都当作学生必须完全一致完成的标准。
10. 必须给出具体分数，不要只给区间。
11. 评价时应先结合教师基础版作品结构、教师基础版运行视频、学生作品结构和学生运行视频，判断学生作品是否实现主要运行效果，再确定完成等级和建议总分。
12. 评分时应采用小学五年级课堂作品评价标准。只要学生作品已经实现本课主要交互和运行目标，即使程序结构不够规范、画面表现一般或缺少拓展功能，也不应过度压低分数。
13. 如果学生作品已经实现教师基础版作品中的主要运行效果，程序能够正常运行，交互基本有效，一般可评为良好完成；如果在此基础上还有较好的拓展或表现效果，可评为优秀完成。
14. 不得将教师基础作品中的变量、广播、倒计时、胜负界面、音效、关卡等具体实现细节自动视为学生必须完成的项目，除非教师上传的评价量表明确要求。
15. 如果学生作品只实现了部分主要运行效果，或运行过程中存在较明显问题，但仍能看出作品基本目标，通常评为基本完成。
16. 如果学生作品只能呈现少量效果，主要交互不完整，或运行逻辑明显不稳定，通常评为部分完成。
17. 如果学生作品无法形成有效作品，主要角色、交互或运行逻辑明显缺失，通常评为未完成。
18. 视频关键帧只能作为运行效果参考；如果视频没有观察到某项功能，但作品结构中可能存在相关程序，不能直接判定为缺失，应写入教师复核点。
19. 不得因为视频未完整呈现某项功能，就直接大幅降低分数；也不得因为画面角色较多或背景完整，就忽略核心交互缺失。
20. 建议总分必须等于各维度建议分数之和。
21. 如果各维度分数相加后的总分与完成等级区间不一致，应调整各维度分数，使建议总分落入对应完成等级区间，并保持评分依据一致。

【完成等级要求】
必须先判断完成等级，再根据完成等级确定建议总分。

建议总分必须落在该完成等级对应的分数区间内：
优秀完成：90—95分
良好完成：85—89分
基本完成：75—84分
部分完成：60—74分
未完成：35—59分

完成等级只能填写以下五个之一：
优秀完成
良好完成
基本完成
部分完成
未完成

不得在完成等级后添加分数区间或解释。

【教师评价案例评分尺度】
{TEACHER_CASE_RULES}

【教师评价案例集】
{teacher_cases_text}

请参考教师评价案例集中评分依据，理解教师评分尺度。
当学生作品与案例中的完成程度相近时，评分应保持一致。
如果案例集与通用评分规则存在差异，应优先参考案例集中体现的评分宽严尺度，但不得改变教师上传评价量表中的评价维度、满分设置和核心评价要求。
案例集只用于学习教师评分尺度，不用于替代教师上传的评价量表。

【输出格式要求】
请严格按照以下格式输出，不要添加额外标题。
字段名前不要添加（一）（二）（三）或1. 2. 3.。
不要输出“维度评分”字段。
各维度评分及依据必须严格使用教师评价量表中的真实维度名称。
核心功能完成情况：
完成等级：
各维度评分及依据：
请逐行输出教师量表中的真实维度名称，例如：
任务完成度评分及依据：36/40。依据：……
编程思维评分及依据：25/30。依据：……
建议总分：
运行画面分析：
总评依据：
改进建议：
教师复核点：

示例1：如果教师量表是“任务完成度40分、编程思维30分、作品表达30分”，则“各维度评分及依据”写成：
任务完成度评分及依据：36/40。依据：……
编程思维评分及依据：25/30。依据：……
作品表达评分及依据：26/30。依据：……

示例2：如果教师量表是“完整性30分、技术性30分、创新性20分、艺术性20分”，则“各维度评分及依据”写成：
完整性评分及依据：28/30。依据：……
技术性评分及依据：26/30。依据：……
创新性评分及依据：17/20。依据：……
艺术性评分及依据：18/20。依据：……

【学生作品结构分析】
{json.dumps(sb3_analysis, ensure_ascii=False, indent=2)}

{reference_text}

{video_text}

【教师上传的作品评价量表】
{rubric_text}
"""

    messages = [
        {"role": "user", "content": prompt}
    ]

    return call_deepseek_full("教师端", messages)

def image_file_to_data_url(uploaded_file):

    """
    把上传的图片转换为 base64 格式，供 OCR 模型读取。
    """
    uploaded_file.seek(0)
    mime_type = uploaded_file.type or "image/png"
    image_bytes = uploaded_file.getvalue()
    encoded_image = base64.b64encode(image_bytes).decode("utf-8")
    return f"data:{mime_type};base64,{encoded_image}"


def recognize_image_text(uploaded_file):
    """
    使用 Qwen-OCR 识别图片中的文字。
    """
    if ocr_client is None:
        return "图片文字识别功能还没有配置 DASHSCOPE_API_KEY，请先在 Streamlit Secrets 中添加 DASHSCOPE_API_KEY。"

    image_url = image_file_to_data_url(uploaded_file)

    try:
        response = ocr_client.chat.completions.create(
            model=QWEN_OCR_MODEL,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": "请提取这张图片中的全部可见文字。尽量保持原有顺序。不要解释，不要扩展，只输出识别到的文字。"
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": image_url
                            }
                        }
                    ]
                }
            ],
            temperature=0,
            max_tokens=1500
        )

        raw_text = response.choices[0].message.content or ""
        return clean_answer(raw_text)

    except Exception as e:
        return f"图片文字识别失败：{e}"




def analyze_student_screenshot_with_qwen_vl(uploaded_file, student_question=""):
    """
    使用 Qwen-VL 读取截图中所有可见的独立脚本，并输出结构化 JSON。

    视觉模型只负责提取证据，不直接判断最终原因。后续程序会逐段分析，
    再结合学生本轮问题进行综合诊断。
    """
    if uploaded_file is None:
        return json.dumps({
            "current_object": "",
            "script_count": 0,
            "scripts": [],
            "other_visible_settings": [],
            "uncertainties": ["本轮未上传截图"],
        }, ensure_ascii=False)

    if ocr_client is None:
        return json.dumps({
            "current_object": "",
            "script_count": 0,
            "scripts": [],
            "other_visible_settings": [],
            "uncertainties": ["未配置DASHSCOPE_API_KEY，无法进行截图结构分析"],
        }, ensure_ascii=False)

    try:
        image_url = image_file_to_data_url(uploaded_file)
        prompt = f"""
你是小学图形化编程截图的“程序结构提取器”，不是答题助手。
请读取截图中所有真正可见的独立脚本，并严格输出一个 JSON 对象。不要输出解释、Markdown 或代码围栏。

学生本轮问题：
{str(student_question or '')[:500]}

提取规则：
1. 一个独立的事件帽积木及其下方连接的积木，算一段脚本。
2. 截图中有几段可见脚本，就必须输出几段；脚本数量不限，不能只输出前两段。
3. 不得补全被遮挡、过小或看不清的积木；不确定的内容写入 uncertainties。
4. 每段脚本必须单独记录，不得把不同脚本合并。
5. 按从上到下的顺序记录积木；保留截图中的角色名、消息名、造型名、变量名和数值。
5.1 必须逐字保留截图中的积木名称和平台用语，例如“重复执行”“循环执行”“重复执行直到”不得互相替换。
6. 对每个积木标记 nesting_depth：事件帽积木为0；事件下直接连接的积木为1；循环或条件内部的积木继续递增。
7. 对条件积木标记 inside_loop；对等待积木标记 input_type，只能是 number、boolean、unclear。
8. modifies 只能从 costume、visibility、position、direction、size、variable、sound、broadcast 中选择。
9. 只提取截图证据，不给修改方案，不判断学生应该怎样做。
10. text 与 input_text 必须使用截图中的原始积木文字，不得把“重复执行”改成“循环执行”，不得替换角色名、造型名、消息名、变量名、运算符或数值。
11. confidence 取0到1；文字、连接或层级不能确认时应降低 confidence，并在 uncertainties 中说明。

JSON 格式：
{{
  "current_object": "当前选中的角色或舞台名称；看不清则为空字符串",
  "script_count": 0,
  "scripts": [
    {{
      "id": "script_1",
      "event": "事件帽积木原文",
      "has_forever_loop": false,
      "blocks": [
        {{
          "order": 1,
          "nesting_depth": 0,
          "text": "积木完整原文，必须逐字保留",
          "block_type": "event|loop|condition|wait|action|other",
          "inside_loop": false,
          "inside_condition": false,
          "input_type": "number|boolean|unclear|none",
          "input_text": "该积木输入槽中可见内容的原文；没有则为空字符串",
          "confidence": 0.0
        }}
      ],
      "modifies": [],
      "uncertainties": []
    }}
  ],
  "other_visible_settings": [],
  "uncertainties": []
}}
"""

        response = ocr_client.chat.completions.create(
            model=QWEN_VL_MODEL,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": image_url}},
                    ],
                }
            ],
            temperature=0,
            max_tokens=2600,
        )

        raw_text = response.choices[0].message.content or ""
        parsed = parse_screenshot_structure(raw_text)
        return json.dumps(parsed, ensure_ascii=False)

    except Exception as exc:
        return json.dumps({
            "current_object": "",
            "script_count": 0,
            "scripts": [],
            "other_visible_settings": [],
            "uncertainties": [f"截图结构分析失败：{exc}"],
        }, ensure_ascii=False)


def safe_bool(value, default=False):
    """可靠解析视觉模型返回的布尔值，避免字符串“false”被当成True。"""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    normalized = str(value or "").strip().lower()
    if normalized in {"true", "1", "yes", "是"}:
        return True
    if normalized in {"false", "0", "no", "否", ""}:
        return False
    return default


def safe_float(value, default=0.0):
    try:
        number = float(value)
    except Exception:
        return default
    return max(0.0, min(1.0, number))


def parse_screenshot_structure(raw_text: str) -> dict:
    """解析并规范视觉模型返回的 JSON；失败时返回可安全处理的空结构。"""
    default = {
        "current_object": "",
        "script_count": 0,
        "scripts": [],
        "other_visible_settings": [],
        "uncertainties": [],
    }
    text = str(raw_text or "").strip()
    if not text:
        default["uncertainties"] = ["视觉模型未返回内容"]
        return default

    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.I)
    text = re.sub(r"\s*```$", "", text)
    try:
        data = json.loads(text)
    except Exception:
        start = text.find("{")
        end = text.rfind("}")
        if start >= 0 and end > start:
            try:
                data = json.loads(text[start:end + 1])
            except Exception:
                default["uncertainties"] = ["截图结构JSON解析失败"]
                return default
        else:
            default["uncertainties"] = ["截图结构JSON解析失败"]
            return default

    if not isinstance(data, dict):
        default["uncertainties"] = ["截图结构不是JSON对象"]
        return default

    scripts = data.get("scripts", [])
    if not isinstance(scripts, list):
        scripts = []

    normalized_scripts = []
    for index, script in enumerate(scripts, start=1):
        if not isinstance(script, dict):
            continue
        blocks = script.get("blocks", [])
        if not isinstance(blocks, list):
            blocks = []
        normalized_blocks = []
        for order, block in enumerate(blocks, start=1):
            if not isinstance(block, dict):
                continue
            normalized_blocks.append({
                "order": int(block.get("order", order) or order),
                "nesting_depth": int(block.get("nesting_depth", 0) or 0),
                "text": str(block.get("text", "") or "").strip(),
                "block_type": str(block.get("block_type", "other") or "other").strip().lower(),
                "inside_loop": safe_bool(block.get("inside_loop", False)),
                "inside_condition": safe_bool(block.get("inside_condition", False)),
                "input_type": str(block.get("input_type", "none") or "none").strip().lower(),
                "input_text": str(block.get("input_text", "") or "").strip(),
                "confidence": safe_float(block.get("confidence", 1.0), 1.0),
            })
        normalized_scripts.append({
            "id": str(script.get("id", f"script_{index}") or f"script_{index}"),
            "event": str(script.get("event", "") or "").strip(),
            "has_forever_loop": safe_bool(script.get("has_forever_loop", False)),
            "blocks": sorted(normalized_blocks, key=lambda item: item["order"]),
            "modifies": [str(x) for x in script.get("modifies", []) if str(x).strip()] if isinstance(script.get("modifies", []), list) else [],
            "uncertainties": [str(x) for x in script.get("uncertainties", []) if str(x).strip()] if isinstance(script.get("uncertainties", []), list) else [],
        })

    reported_count = data.get("script_count", len(normalized_scripts))
    try:
        reported_count = int(reported_count)
    except Exception:
        reported_count = len(normalized_scripts)

    uncertainties = data.get("uncertainties", [])
    if not isinstance(uncertainties, list):
        uncertainties = [str(uncertainties)] if uncertainties else []
    if reported_count != len(normalized_scripts):
        uncertainties.append(
            f"视觉模型报告{reported_count}段脚本，但实际解析到{len(normalized_scripts)}段"
        )

    return {
        "current_object": str(data.get("current_object", "") or "").strip(),
        "script_count": len(normalized_scripts),
        "scripts": normalized_scripts,
        "other_visible_settings": data.get("other_visible_settings", []) if isinstance(data.get("other_visible_settings", []), list) else [],
        "uncertainties": [str(x) for x in uncertainties if str(x).strip()],
    }


def get_screenshot_structure(screenshot_analysis) -> dict:
    """兼容 JSON 字符串或字典形式的截图结构。"""
    if isinstance(screenshot_analysis, dict):
        return screenshot_analysis
    return parse_screenshot_structure(str(screenshot_analysis or ""))


def script_text(script: dict) -> str:
    parts = [str(script.get("event", "") or "")]
    parts.extend(str(block.get("text", "") or "") for block in script.get("blocks", []))
    return "\n".join(part for part in parts if part).strip()


def script_compact(script: dict) -> str:
    return re.sub(r"\s+", "", script_text(script))


def script_has_loop(script: dict) -> bool:
    if script.get("has_forever_loop"):
        return True
    for block in script.get("blocks", []):
        text = re.sub(r"\s+", "", str(block.get("text", "") or ""))
        if block.get("block_type") == "loop" or any(word in text for word in ["循环执行", "重复执行", "重复执行直到"]):
            return True
    return False


def find_relevant_hit_scripts(structure: dict):
    """在所有脚本中查找包含鼠标与碰撞判断的脚本。"""
    results = []
    for script in structure.get("scripts", []):
        compact = script_compact(script)
        has_mouse = any(term in compact for term in ["按下鼠标", "鼠标按下", "鼠标键按下", "按下鼠标键"])
        has_touch = "碰到" in compact
        has_condition = any(
            block.get("block_type") == "condition" or "如果" in str(block.get("text", ""))
            for block in script.get("blocks", [])
        )
        if has_mouse and has_touch and has_condition:
            condition_blocks = [
                block for block in script.get("blocks", [])
                if block.get("block_type") == "condition" or "如果" in str(block.get("text", ""))
            ]
            condition_inside_loop = any(bool(block.get("inside_loop")) for block in condition_blocks)
            results.append({
                "script": script,
                "condition_inside_loop": condition_inside_loop,
                "has_loop": script_has_loop(script),
            })
    return results


BOOLEAN_WAIT_CONDITION_MARKERS = [
    "按下鼠标", "碰到", "碰到颜色", "颜色碰到", "按下", "键",
    "鼠标指针", "边缘", "询问", "回答", "计时器", "音量", "距离",
]
BOOLEAN_WAIT_OPERATOR_MARKERS = ["与", "且", "或", "不成立", "=", ">", "<", "？", "?"]


def wait_seconds_slot_text(block: dict) -> str:
    """提取“等待（）秒”数值槽里的文字，优先使用视觉模型给出的 input_text。"""
    input_text = str(block.get("input_text", "") or "").strip()
    if input_text:
        return input_text
    text = str(block.get("text", "") or "").strip()
    match = re.search(r"等待\s*(.*?)\s*秒", text)
    return match.group(1).strip(" <>[]（）()") if match else ""


def is_wait_seconds_block(block: dict) -> bool:
    text = re.sub(r"\s+", "", str(block.get("text", "") or ""))
    return block.get("block_type") == "wait" or ("等待" in text and "秒" in text)


def text_looks_like_boolean_condition(text: str) -> bool:
    compact = re.sub(r"\s+", "", str(text or ""))
    if not compact:
        return False
    has_condition_marker = any(marker in compact for marker in BOOLEAN_WAIT_CONDITION_MARKERS)
    has_operator_marker = any(marker in compact for marker in BOOLEAN_WAIT_OPERATOR_MARKERS)
    return has_condition_marker or has_operator_marker


def find_boolean_wait_errors(structure: dict):
    """查找把布尔条件放入‘等待若干秒’数值输入框的脚本。"""
    errors = []
    for script in structure.get("scripts", []):
        script_text_compact = re.sub(
            r"\s+", "", " ".join(str(block.get("text", "") or "") for block in script.get("blocks", []))
        )
        script_has_condition_text = text_looks_like_boolean_condition(script_text_compact) or "如果" in script_text_compact
        for block in script.get("blocks", []):
            text = re.sub(r"\s+", "", str(block.get("text", "") or ""))
            slot_text = wait_seconds_slot_text(block)
            is_wait_block = is_wait_seconds_block(block)
            if is_wait_block and block.get("input_type") == "boolean":
                errors.append({"script": script, "block": block})
            elif is_wait_block and text_looks_like_boolean_condition(slot_text or text):
                errors.append({"script": script, "block": block})
            elif is_wait_block and not re.search(r"\d", text) and script_has_condition_text:
                errors.append({"script": script, "block": block})
    return errors


def find_parallel_modification_conflicts(structure: dict):
    """综合所有脚本，查找两段及以上持续脚本同时修改同一属性的情况。"""
    property_to_scripts = {}
    for script in structure.get("scripts", []):
        if not script_has_loop(script):
            continue
        for prop in script.get("modifies", []):
            property_to_scripts.setdefault(prop, []).append(script)
    return {
        prop: scripts
        for prop, scripts in property_to_scripts.items()
        if len(scripts) >= 2
    }


def classify_student_debug_intent(student_question: str) -> str:
    """根据学生本轮文字问题判断当前最需要解释的运行现象。"""
    question = re.sub(r"\s+", "", str(student_question or ""))

    if any(word in question for word in ["闪", "闪现", "不停切换", "一直切换"]):
        return "flicker"
    if (
        any(word in question for word in ["打到", "击中", "点击", "碰到"])
        and any(word in question for word in ["不哭", "没有哭", "没哭", "不变造型", "没反应", "没有反应"])
    ):
        return "hit_feedback_missing"
    if any(word in question for word in ["不加分", "分数不变", "得分不增加", "没有加分"]):
        return "score_not_increasing"
    if any(word in question for word in ["只执行一次", "只动一次", "只出现一次"]):
        return "runs_once"
    if any(word in question for word in ["一直加分", "分数一直增加", "连续加分"]):
        return "score_repeats"
    if any(word in question for word in ["不动", "没有移动", "不跟随"]):
        return "not_moving"
    return "general"


def extract_hit_target_from_structure(structure: dict, ocr_text: str = "") -> str:
    combined = "\n".join(script_text(script) for script in structure.get("scripts", [])) + "\n" + str(ocr_text or "")
    compact = re.sub(r"\s+", "", combined)
    patterns = [
        r"碰到[“\"']?([^？?<>〈〉（）()\s]{1,12})[”\"']?[？?]",
    ]
    for pattern in patterns:
        match = re.search(pattern, compact)
        if match:
            target = match.group(1).strip("？?<>〈〉")
            if target and target not in ["鼠标指针", "边缘"]:
                return target
    for candidate in ["地鼠锤子", "锤子", "捕虫网", "小猫", "老鼠", "电池", "星星", "宝箱"]:
        if candidate in compact:
            return candidate
    return "目标角色"


def extract_costume_sequence(script: dict):
    names = []
    for block in script.get("blocks", []):
        text = str(block.get("text", "") or "")
        match = re.search(r"换成[“\"']?(.+?)[”\"']?造型", text)
        if match:
            name = match.group(1).strip(" “\"'”")
            if name and name not in names:
                names.append(name)
    return names


def extract_wait_seconds(script: dict, default="0.5") -> str:
    for block in script.get("blocks", []):
        if block.get("block_type") == "wait" and block.get("input_type") == "number":
            match = re.search(r"等待\s*([0-9]+(?:\.[0-9]+)?)\s*秒", str(block.get("text", "")))
            if match:
                return match.group(1)
    return default


def extract_exact_block_label(script: dict, block_type: str, candidates=None, default="") -> str:
    """从截图结构中提取学生实际使用的积木原文。"""
    candidates = candidates or []
    for block in script.get("blocks", []):
        text = str(block.get("text", "") or "").strip()
        compact = re.sub(r"\s+", "", text)
        if block.get("block_type") == block_type:
            if candidates:
                for candidate in candidates:
                    if candidate in compact:
                        return candidate
            if text:
                return text
        for candidate in candidates:
            if candidate in compact:
                return candidate
    return default


def infer_loop_label(script: dict, structure: dict = None) -> str:
    """优先使用当前脚本或同一截图中实际出现的持续循环积木名称。"""
    labels = ["重复执行直到", "重复执行", "循环执行"]
    label = extract_exact_block_label(script, "loop", labels, "")
    if label and label != "重复执行直到":
        return label
    if structure:
        for other in structure.get("scripts", []):
            label = extract_exact_block_label(other, "loop", labels, "")
            if label and label != "重复执行直到":
                return label
    # 截图中缺少该类积木时，只能使用当前平台最常见的标准名称。
    return "循环执行"


def get_event_text(script: dict) -> str:
    event = str(script.get("event", "") or "").strip()
    if event:
        return event
    for block in script.get("blocks", []):
        if block.get("block_type") == "event":
            return str(block.get("text", "") or "").strip()
    return ""


def get_non_event_blocks(script: dict):
    """返回脚本中除事件帽积木外的全部积木，保持截图原顺序。"""
    event = re.sub(r"\s+", "", get_event_text(script))
    result = []
    event_skipped = False
    for block in sorted(script.get("blocks", []), key=lambda item: item.get("order", 0)):
        text = str(block.get("text", "") or "").strip()
        compact = re.sub(r"\s+", "", text)
        if block.get("block_type") == "event" or (event and compact == event and not event_skipped):
            event_skipped = True
            continue
        if text:
            result.append(block)
    return result


def block_indent(block: dict, base_shift: int = 0) -> str:
    depth = max(1, int(block.get("nesting_depth", 1) or 1) + base_shift)
    return "    " * max(0, depth - 1)


def render_script_exact(script: dict, insert_loop_label: str = "") -> str:
    """
    使用视觉模型提取的积木原文重建局部脚本。
    除必要新增的循环外，不重写任何已有积木名称、参数或对象名称。
    """
    lines = []
    event = get_event_text(script)
    if event:
        lines.append(event)
    blocks = get_non_event_blocks(script)
    if insert_loop_label:
        lines.append(insert_loop_label)
        for block in blocks:
            lines.append("    " + block_indent(block) + str(block.get("text", "") or "").strip())
    else:
        for block in blocks:
            lines.append(block_indent(block) + str(block.get("text", "") or "").strip())
    return "\n".join(line for line in lines if line.strip())


def extract_boolean_input_text(block: dict) -> str:
    """提取被错误放入数值槽中的布尔条件原文。"""
    return wait_seconds_slot_text(block)


def infer_condition_wrapper(structure: dict, condition_text: str) -> str:
    """优先参照同一截图中真实出现的条件积木写法；没有时使用平台标准写法。"""
    for script in structure.get("scripts", []):
        for block in script.get("blocks", []):
            if block.get("block_type") != "condition":
                continue
            raw = str(block.get("text", "") or "").strip()
            compact = re.sub(r"\s+", "", raw)
            if "如果" in compact and "那么" in compact:
                return f"如果 <{condition_text}> 那么"
    return f"如果 <{condition_text}> 那么"


def infer_numeric_wait_text(structure: dict, preferred_script: dict, default_seconds="0.5") -> str:
    """优先复制截图中已出现的数值等待积木原文；没有时才新增标准等待积木。"""
    scripts = [preferred_script] + [s for s in structure.get("scripts", []) if s is not preferred_script]
    for script in scripts:
        for block in script.get("blocks", []):
            if block.get("block_type") == "wait" and block.get("input_type") == "number":
                raw = str(block.get("text", "") or "").strip()
                if raw:
                    return raw
    return f"等待{default_seconds}秒"


def _costume_name_from_block(block: dict) -> str:
    text = str(block.get("text", "") or "").strip()
    match = re.search(r"换成\s*[“\"']?(.+?)[”\"']?\s*造型", text)
    return match.group(1).strip(" ‘\"'”") if match else ""


def _select_feedback_and_reset_costumes(blocks: list, student_question: str = ""):
    """
    依据学生提问与造型语义选择“击中反馈造型”和“恢复造型”。
    只有证据明确时才调整顺序，避免机械保留错误顺序。
    """
    costume_blocks = [b for b in blocks if _costume_name_from_block(b)]
    if len(costume_blocks) < 2:
        return None, None

    question = re.sub(r"\s+", "", str(student_question or ""))
    feedback_keywords = ["哭", "受伤", "疼", "击中", "被打", "爆炸", "倒下", "死亡", "失败", "消失"]
    reset_keywords = ["洞", "正常", "默认", "原始", "站立", "等待", "初始"]

    def feedback_score(block):
        name = _costume_name_from_block(block)
        score = 0
        if name and name in question:
            score += 20
        for kw in feedback_keywords:
            if kw in question and kw in name:
                score += 12
            elif kw in name:
                score += 4
        if any(expr in question for expr in ["没有哭", "没哭", "不会哭", "不哭"]) and "哭" in name:
            score += 20
        return score

    def reset_score(block):
        name = _costume_name_from_block(block)
        score = 0
        for kw in reset_keywords:
            if kw in name:
                score += 8
        return score

    feedback = max(costume_blocks, key=feedback_score)
    remaining = [b for b in costume_blocks if b is not feedback]
    reset = max(remaining, key=reset_score) if remaining else None

    if feedback_score(feedback) <= 0 or reset is None:
        return None, None
    return feedback, reset




def _condition_blocks_with_scripts(structure: dict):
    """返回所有条件积木及其所属脚本，保持截图顺序。"""
    items = []
    for script in structure.get("scripts", []):
        for block in script.get("blocks", []):
            raw = str(block.get("text", "") or "").strip()
            if block.get("block_type") == "condition" or "如果" in raw:
                items.append({"script": script, "block": block, "text": raw})
    return items


def _student_expects_mouse_press(question: str) -> bool:
    """判断学生描述的操作是否明确包含鼠标点击/按下。"""
    q = re.sub(r"\s+", "", str(question or ""))
    return any(term in q for term in [
        "点击", "点到", "点一下", "按下鼠标", "按鼠标", "鼠标点击", "用鼠标打", "打到"
    ])


def _replace_mouse_pointer_touch_with_press(condition_text: str) -> str:
    """只替换条件中的‘碰到鼠标指针？’，保留其他原文。"""
    raw = str(condition_text or "")
    patterns = [
        r"碰到\s*[“\"']?鼠标指针[”\"']?\s*[？?]?",
        r"碰到\s*鼠标指针\s*[？?]?",
    ]
    new = raw
    for pattern in patterns:
        new = re.sub(pattern, "按下鼠标？", new)
    return new


def find_mouse_trigger_mismatch(structure: dict, student_question: str):
    """
    查找“学生要用鼠标点击/按下，但条件写成碰到鼠标指针”的明确不一致。
    仅在同一条件还包含另一个碰撞目标时触发，避免误改真正需要鼠标指针碰撞的任务。
    """
    if not _student_expects_mouse_press(student_question):
        return []
    results = []
    for item in _condition_blocks_with_scripts(structure):
        compact = re.sub(r"\s+", "", item["text"])
        has_pointer_touch = "碰到鼠标指针" in compact
        touch_targets = re.findall(r"碰到[“\"']?([^？?<>〈〉（）()\s]{1,16})[”\"']?[？?]?", compact)
        non_pointer_targets = [t for t in touch_targets if t not in {"鼠标指针", "边缘"}]
        if has_pointer_touch and non_pointer_targets:
            block = item["block"]
            confidence = safe_float(block.get("confidence", 1.0), 1.0)
            if confidence >= 0.75:
                results.append({
                    **item,
                    "target": non_pointer_targets[0],
                    "replacement_condition": _replace_mouse_pointer_touch_with_press(item["text"]),
                })
    return results


def _question_expected_costume_keyword(student_question: str) -> str:
    q = re.sub(r"\s+", "", str(student_question or ""))
    for keyword in ["哭", "笑", "消失", "受伤", "倒下", "爆炸"]:
        if keyword in q:
            return keyword
    return ""


def find_expected_feedback_mismatch(structure: dict, student_question: str):
    """检查条件体内的反馈动作是否与学生明确说出的期望效果一致。"""
    expected = _question_expected_costume_keyword(student_question)
    if not expected:
        return []
    results = []
    for item in _condition_blocks_with_scripts(structure):
        script = item["script"]
        condition = item["block"]
        cond_order = int(condition.get("order", 0) or 0)
        cond_depth = int(condition.get("nesting_depth", 1) or 1)
        body = []
        for block in script.get("blocks", []):
            order = int(block.get("order", 0) or 0)
            depth = int(block.get("nesting_depth", 1) or 1)
            if order > cond_order and (block.get("inside_condition") or depth > cond_depth):
                body.append(block)
        costume_names = [_costume_name_from_block(b) for b in body]
        costume_names = [n for n in costume_names if n]
        if costume_names and not any(expected in n for n in costume_names):
            results.append({
                **item,
                "expected": expected,
                "actual_costumes": costume_names,
            })
    return results


def render_script_with_condition_replacement(script: dict, condition_block: dict, new_condition_text: str) -> str:
    """重建相关局部脚本，只替换一个条件积木，其他已有积木逐字保留。"""
    lines = []
    event = get_event_text(script)
    if event:
        lines.append(event)
    target_order = int(condition_block.get("order", 0) or 0)
    for block in get_non_event_blocks(script):
        raw = str(block.get("text", "") or "").strip()
        if int(block.get("order", 0) or 0) == target_order:
            raw = new_condition_text
        lines.append(block_indent(block) + raw)
    return "\n".join(line for line in lines if line.strip())


def build_mouse_trigger_repair(structure: dict, mismatch: dict) -> str:
    """保留原脚本，只把‘碰到鼠标指针？’改为‘按下鼠标？’。"""
    return render_script_with_condition_replacement(
        mismatch["script"],
        mismatch["block"],
        mismatch["replacement_condition"],
    )

def build_boolean_wait_repair(
    structure: dict,
    script: dict,
    bad_block: dict,
    student_question: str = "",
) -> str:
    """
    将错误的布尔等待槽改成“条件积木 + 数值等待”。
    已有积木名称按截图原文保留，但会结合学生问题与正常运行逻辑，
    将明确的反馈造型放在条件成立后，将恢复造型放在等待之后。
    """
    event = get_event_text(script)
    loop_label = infer_loop_label(script, structure)
    condition_text = extract_boolean_input_text(bad_block)
    if not condition_text:
        return ""
    condition_line = infer_condition_wrapper(structure, condition_text)
    wait_line = infer_numeric_wait_text(structure, script)

    blocks = get_non_event_blocks(script)
    bad_order = int(bad_block.get("order", 0) or 0)
    remaining_blocks = [
        block for block in blocks
        if not (block is bad_block or int(block.get("order", 0) or 0) == bad_order)
        and block.get("block_type") != "loop"
    ]

    feedback_block, reset_block = _select_feedback_and_reset_costumes(
        remaining_blocks, student_question
    )

    lines = [event] if event else []
    if script_has_loop(script):
        existing_loop = extract_exact_block_label(script, "loop", default=loop_label)
        lines.append(existing_loop or loop_label)
    else:
        lines.append(loop_label)
    lines.append("    " + condition_line)

    if feedback_block and reset_block:
        # 明确的语义顺序：条件成立 → 反馈造型 → 等待 → 恢复造型。
        lines.append("        " + str(feedback_block.get("text", "") or "").strip())
        lines.append("        " + wait_line)
        lines.append("        " + str(reset_block.get("text", "") or "").strip())
        for block in remaining_blocks:
            if block is feedback_block or block is reset_block:
                continue
            raw = str(block.get("text", "") or "").strip()
            if raw:
                lines.append("        " + raw)
    else:
        # 证据不足时不擅自改变动作含义，只修正条件与等待的结构。
        before = []
        after = []
        for block in remaining_blocks:
            raw = str(block.get("text", "") or "").strip()
            if int(block.get("order", 0) or 0) < bad_order:
                before.append(raw)
            else:
                after.append(raw)
        for raw in before:
            lines.append("        " + raw)
        lines.append("        " + wait_line)
        for raw in after:
            lines.append("        " + raw)

    return "\n".join(line for line in lines if line.strip())


def build_missing_loop_repair(structure: dict, script: dict) -> str:
    """
    让相关条件进入持续检测。
    没有循环时只新增一个循环；已有循环但条件在循环外时，只调整原有积木的嵌套层级。
    """
    loop_label = infer_loop_label(script, structure)
    if not script_has_loop(script):
        return render_script_exact(script, insert_loop_label=loop_label)

    event = get_event_text(script)
    blocks = get_non_event_blocks(script)
    loop_indexes = [
        index for index, block in enumerate(blocks)
        if block.get("block_type") == "loop" or any(
            word in re.sub(r"\s+", "", str(block.get("text", "") or ""))
            for word in ["循环执行", "重复执行", "重复执行直到"]
        )
    ]
    condition_indexes = [
        index for index, block in enumerate(blocks)
        if block.get("block_type") == "condition" or "如果" in str(block.get("text", "") or "")
    ]
    if not loop_indexes or not condition_indexes:
        return render_script_exact(script)

    loop_index = loop_indexes[0]
    condition_index = condition_indexes[0]
    lines = [event] if event else []
    loop_depth = max(1, int(blocks[loop_index].get("nesting_depth", 1) or 1))
    condition_order = int(blocks[condition_index].get("order", condition_index + 1) or condition_index + 1)

    for index, block in enumerate(blocks):
        raw = str(block.get("text", "") or "").strip()
        depth = max(1, int(block.get("nesting_depth", 1) or 1))
        # 条件及其后续同级动作移动到已有循环内部；已有更深层级继续保留相对嵌套。
        if int(block.get("order", index + 1) or index + 1) >= condition_order and depth <= loop_depth:
            depth = loop_depth + 1
        lines.append("    " * max(0, depth - 1) + raw)
    return "\n".join(line for line in lines if line.strip())


def build_screenshot_hard_constraints(
    screenshot_analysis,
    ocr_text: str,
    student_question: str = "",
):
    """
    逐段分析全部脚本，并严格按以下优先级诊断：
    1. 学生操作方式与条件积木是否一致；
    2. 条件后的反馈动作是否符合学生目标；
    3. 条件是否持续检测；
    4. 条件或输入槽结构是否错误；
    5. 最后才考虑多脚本覆盖冲突。
    """
    structure = get_screenshot_structure(screenshot_analysis)
    intent = classify_student_debug_intent(student_question)
    question_goal = infer_question_goal(student_question)
    explicit_unrelated_goal = bool(
        question_goal.get("actions")
        and not any(action in {"切换受击造型", "改变分数"} for action in question_goal.get("actions", []))
    )

    trigger_mismatches = find_mouse_trigger_mismatch(structure, student_question)
    feedback_mismatches = find_expected_feedback_mismatch(structure, student_question)
    hit_scripts = find_relevant_hit_scripts(structure)
    boolean_wait_errors = find_boolean_wait_errors(structure)
    parallel_conflicts = find_parallel_modification_conflicts(structure)

    evidence = []
    for item in trigger_mismatches:
        evidence.append(
            f"{item['script'].get('id')}条件使用‘碰到鼠标指针’，但学生描述的是鼠标点击/按下；另一个碰撞目标为{item.get('target', '')}"
        )
    for item in feedback_mismatches:
        evidence.append(
            f"{item['script'].get('id')}条件体中的造型为{item.get('actual_costumes', [])}，未包含学生期望的‘{item.get('expected', '')}’效果"
        )
    for item in hit_scripts:
        script = item["script"]
        evidence.append(
            f"{script.get('id')}含鼠标与碰撞条件；条件在循环中={item['condition_inside_loop']}；脚本含循环={item['has_loop']}"
        )
    for item in boolean_wait_errors:
        evidence.append(f"{item['script'].get('id')}把布尔条件放入等待秒数输入框")
    for prop, scripts in parallel_conflicts.items():
        evidence.append(f"{len(scripts)}段持续脚本同时修改{prop}")

    direct_answer = ""
    primary_hit = hit_scripts[0] if hit_scripts else None

    # 最高优先级：学生说“点击/按下鼠标”，但条件却写成“碰到鼠标指针”。
    if trigger_mismatches:
        mismatch = trigger_mismatches[0]
        program = build_mouse_trigger_repair(structure, mismatch)
        original_condition = mismatch.get("text", "")
        target = mismatch.get("target", "目标角色")
        if program:
            direct_answer = (
                f"条件中的“碰到鼠标指针？”和你的操作方式不一致。你是按下鼠标并让锤子碰到地鼠，所以把“碰到鼠标指针？”换成“按下鼠标？”，保留“碰到{target}？”。\n\n"
                f"修改后像这样：\n[[PROGRAM]]\n{program}\n[[/PROGRAM]]\n"
                "改好后按下鼠标再用锤子碰到地鼠，看看反馈是否出现。"
            )

    # 第二优先级：条件成立后的动作本身就不是学生要的效果。
    elif feedback_mismatches:
        mismatch = feedback_mismatches[0]
        actual = "、".join(mismatch.get("actual_costumes", []))
        expected = mismatch.get("expected", "")
        direct_answer = (
            f"这段条件成立后切换的是“{actual}”造型，不是你想要的“{expected}”效果。"
            f"先把条件里面对应的造型积木改成包含“{expected}”的造型，其他脚本先不要删。"
        )

    # 第三优先级：相关条件没有持续检测。
    elif intent in {"hit_feedback_missing", "score_not_increasing", "runs_once"} and primary_hit and not primary_hit["condition_inside_loop"]:
        script = primary_hit["script"]
        program = build_missing_loop_repair(structure, script)
        event = get_event_text(script) or "这个事件"
        if intent == "hit_feedback_missing":
            reason = f"“{event}”触发后，这个条件只判断一次，所以之后再击中目标时不会重新检查。"
            verification = "改好后再击中一次，看看造型会不会变化。"
        elif intent == "score_not_increasing":
            reason = f"“{event}”触发后，这个条件只判断一次，所以之后再击中目标时不会重新检查加分条件。"
            verification = "改好后再击中一次，看看分数会不会增加。"
        else:
            reason = "这段条件只判断一次，所以后续操作不会继续触发。"
            verification = "改好后运行一下，看看程序是否会持续判断。"
        if program:
            direct_answer = f"{reason}\n\n修改后像这样：\n[[PROGRAM]]\n{program}\n[[/PROGRAM]]\n{verification}"

    # 第四优先级：布尔条件被错误放入数值等待槽。
    elif boolean_wait_errors and not explicit_unrelated_goal:
        err = boolean_wait_errors[0]
        program = build_boolean_wait_repair(structure, err["script"], err["block"], student_question)
        if program:
            condition_text = extract_boolean_input_text(err["block"])
            direct_answer = (
                f"“{condition_text}”被放进了“等待（）秒”的数值框，等待积木这里应该填写时间。"
                "把这个条件移到“如果……那么……”中，并把等待改成数值等待。\n\n"
                f"修改后像这样：\n[[PROGRAM]]\n{program}\n[[/PROGRAM]]\n"
                "改好后运行一下，看看效果是否正常。"
            )

    # 最后才考虑多脚本冲突，而且仅在前面没有更直接错误时使用。
    elif intent == "flicker" and "costume" in parallel_conflicts:
        direct_answer = (
            "截图中有两段以上的持续脚本同时修改同一个角色的造型，后执行的造型会覆盖前一段，所以看起来会不停闪。"
            "先找出两段脚本中重复控制造型的部分，再只调整与当前效果直接相关的那一处，不要删除整段脚本。"
        )

    constraints = {
        "intent": intent,
        "current_object": structure.get("current_object", ""),
        "script_count": structure.get("script_count", 0),
        "all_scripts_analyzed": True,
        "diagnosis_priority": [
            "操作方式与条件是否一致",
            "反馈动作是否符合学生目标",
            "条件是否持续检测",
            "积木槽位与嵌套是否正确",
            "多脚本冲突",
        ],
        "evidence": evidence,
        "uncertainties": structure.get("uncertainties", []),
        "rules": [
            "必须先逐段分析所有可见脚本，再综合脚本之间的关系。",
            "必须先比较学生描述的操作方式与条件积木，不得把‘碰到鼠标指针’当成‘按下鼠标’。",
            "学生明确要求某个反馈效果时，必须先检查条件体内是否存在对应动作。",
            "只有触发条件、反馈动作和持续检测都正确时，才把多脚本冲突作为主要原因。",
            "程序框中的已有积木必须逐字复制截图结构中的text，不得重新命名。",
            "修复时优先只替换一个错误积木或调整一个嵌套关系，不得删除无关整段脚本。",
            "不得无依据增加停止其他脚本、状态变量或复杂控制方案。",
        ],
    }
    return json.dumps(constraints, ensure_ascii=False), direct_answer


# =========================
# 4.5.1 连续截图调试上下文
# =========================

def _compact_text(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def extract_structure_entities(structure: dict) -> dict:
    """从截图结构中提取可用于新旧截图关系判断的稳定实体。"""
    entities = {
        "object": str(structure.get("current_object", "") or "").strip(),
        "events": set(),
        "messages": set(),
        "costumes": set(),
        "variables": set(),
        "targets": set(),
        "block_terms": set(),
    }
    for script in structure.get("scripts", []):
        event = str(script.get("event", "") or "").strip()
        if event:
            entities["events"].add(event)
        for block in script.get("blocks", []):
            text = str(block.get("text", "") or "").strip()
            if not text:
                continue
            compact = _compact_text(text)
            entities["block_terms"].add(compact)
            for pattern, key in [
                (r"(?:广播|当接收到)\s*[“\"']?(.+?)[”\"']?(?:$|并等待|$)", "messages"),
                (r"换成\s*[“\"']?(.+?)[”\"']?\s*造型", "costumes"),
                (r"碰到\s*[“\"']?(.+?)[”\"']?\s*\?", "targets"),
                (r"(?:将|把)\s*[“\"']?(.+?)[”\"']?\s*(?:设为|增加|改变)", "variables"),
            ]:
                match = re.search(pattern, text)
                if match:
                    entities[key].add(match.group(1).strip())
    return entities


def structure_summary_for_prompt(structure: dict, max_scripts: int = 12) -> str:
    """生成可安全传给回答模型的截图结构摘要。"""
    lines = [
        f"当前对象：{structure.get('current_object', '') or '未确认'}",
        f"可见脚本数：{structure.get('script_count', 0)}",
    ]
    for index, script in enumerate(structure.get("scripts", [])[:max_scripts], start=1):
        lines.append(f"脚本{index}：")
        for block in script.get("blocks", []):
            raw = str(block.get("text", "") or "").strip()
            if not raw:
                continue
            depth = max(0, int(block.get("nesting_depth", 0) or 0))
            lines.append("  " * depth + raw)
    uncertainties = structure.get("uncertainties", [])
    if uncertainties:
        lines.append("不确定项：" + "；".join(map(str, uncertainties[:4])))
    return "\n".join(lines)


def infer_question_goal(question: str) -> dict:
    """从学生原话中提取触发方式和期望动作，避免用常见场景替代学生目标。"""
    text = _compact_text(question)
    trigger = ""
    if any(k in text for k in ["点击角色", "点角色", "点击锤子", "点锤子", "鼠标点击"]):
        trigger = "点击角色"
    elif any(k in text for k in ["碰到", "撞到", "接触"]):
        trigger = "碰撞"
    elif any(k in text for k in ["按键", "空格键", "方向键", "键盘"]):
        trigger = "按键"
    elif any(k in text for k in ["绿旗", "开始"]):
        trigger = "开始事件"

    actions = []
    action_map = [
        (["左转"], "左转"), (["右转"], "右转"), (["旋转", "转动", "挥动", "摆动"], "转动"),
        (["跟随鼠标", "移到鼠标", "鼠标指针"], "跟随鼠标"),
        (["哭", "受伤"], "切换受击造型"), (["加分", "得分"], "改变分数"),
        (["隐藏", "消失"], "隐藏"), (["显示", "出现"], "显示"),
        (["播放声音", "声音"], "播放声音"),
    ]
    for keys, label in action_map:
        if any(k in text for k in keys) and label not in actions:
            actions.append(label)

    return {"trigger": trigger, "actions": actions, "raw": str(question or "").strip()}


def is_followup_or_correction(question: str) -> bool:
    text = _compact_text(question)
    if not text:
        return False
    cues = [
        "我的意思", "我是说", "不是", "是我", "对", "不对", "还是", "继续", "刚才", "上面", "这个",
        "它", "那里", "这样", "只要", "我想要", "我希望", "应该是", "改成",
    ]
    return len(text) <= 40 or any(cue in text for cue in cues)


def is_text_related_to_context(question: str, context: dict) -> bool:
    """判断无新截图的文字是否仍在追问最近一次截图。"""
    if not context:
        return False
    if is_followup_or_correction(question):
        return True
    text = _compact_text(question)
    if not text:
        return False
    structure = context.get("structure", {}) or {}
    entities = extract_structure_entities(structure)
    candidates = [entities.get("object", "")]
    for key in ["messages", "costumes", "variables", "targets"]:
        candidates.extend(list(entities.get(key, set()) or set()))
    return any(_compact_text(item) and _compact_text(item) in text for item in candidates)


def _set_overlap(a, b) -> float:
    a = set(a or [])
    b = set(b or [])
    if not a or not b:
        return 0.0
    return len(a & b) / max(1, len(a | b))


def compare_debug_contexts(previous: dict, new_structure: dict, new_question: str) -> dict:
    """判断新截图是同一任务、同一作品的新任务、无关内容，还是证据不足。"""
    if not previous:
        return {
            "relation": "new_context",
            "reason": "没有可比较的上一轮截图上下文",
            "inherit_project_context": False,
            "inherit_debug_diagnosis": False,
        }

    prev_structure = previous.get("structure", {}) or {}
    prev_entities = extract_structure_entities(prev_structure)
    new_entities = extract_structure_entities(new_structure)
    prev_obj = _compact_text(prev_entities.get("object", ""))
    new_obj = _compact_text(new_entities.get("object", ""))

    same_object = bool(prev_obj and new_obj and prev_obj == new_obj)
    event_overlap = _set_overlap(prev_entities["events"], new_entities["events"])
    message_overlap = _set_overlap(prev_entities["messages"], new_entities["messages"])
    costume_overlap = _set_overlap(prev_entities["costumes"], new_entities["costumes"])
    target_overlap = _set_overlap(prev_entities["targets"], new_entities["targets"])
    term_overlap = _set_overlap(prev_entities["block_terms"], new_entities["block_terms"])
    project_score = max(message_overlap, target_overlap, costume_overlap, event_overlap)
    task_score = (0.45 if same_object else 0.0) + 0.35 * term_overlap + 0.20 * event_overlap

    prev_goal = previous.get("goal", {}) or {}
    new_goal = infer_question_goal(new_question)
    trigger_changed = bool(prev_goal.get("trigger") and new_goal.get("trigger") and prev_goal.get("trigger") != new_goal.get("trigger"))
    action_changed = bool(prev_goal.get("actions") and new_goal.get("actions") and set(prev_goal.get("actions", [])) != set(new_goal.get("actions", [])))

    if same_object and task_score >= 0.55 and not (trigger_changed or action_changed):
        relation = "same_task"
        reason = "当前角色和主要脚本高度重合，且学生目标未发生明显变化"
        inherit_debug = True
    elif same_object or project_score >= 0.25:
        relation = "same_project_new_task"
        reason = "截图仍属于同一作品或共享相同消息、造型、碰撞对象，但当前角色或目标发生变化"
        inherit_debug = False
    elif prev_obj and new_obj and prev_obj != new_obj and project_score == 0:
        relation = "unrelated"
        reason = "角色和关键脚本实体均未发现可靠重合"
        inherit_debug = False
    else:
        relation = "uncertain"
        reason = "截图中的角色或关键脚本信息不足，无法可靠判断与上一轮的关系"
        inherit_debug = False

    return {
        "relation": relation,
        "reason": reason,
        "inherit_project_context": relation in {"same_task", "same_project_new_task"},
        "inherit_debug_diagnosis": inherit_debug,
        "scores": {
            "task": round(task_score, 3),
            "project": round(project_score, 3),
            "term_overlap": round(term_overlap, 3),
        },
    }


def build_active_debug_context_prompt(context: dict) -> str:
    if not context:
        return "当前没有可继承的截图调试上下文。"
    structure = context.get("structure", {}) or {}
    goal = context.get("goal", {}) or {}
    relation = context.get("relation", {}) or {}
    return f"""
【最近一次有效截图调试上下文】
学生当时的问题：{context.get('question', '') or '未说明'}
学生目标：触发方式={goal.get('trigger', '') or '未确认'}；期望动作={'、'.join(goal.get('actions', [])) or '未确认'}
新旧截图关系：{relation.get('relation', 'unknown')}；{relation.get('reason', '')}
截图结构摘要：
{structure_summary_for_prompt(structure)}

使用规则：
1. 这段上下文只用于理解学生对最近截图的补充、纠正或追问。
2. 学生当前原话优先级最高；若当前原话修正了旧目标，必须更新理解，不得坚持旧答案。
3. 不得把旧截图中的角色、诊断或程序移植到明显不同的新任务中。
""".strip()


def update_project_context(old_project: dict, new_structure: dict, relation: dict) -> dict:
    """仅合并作品层面的公共实体，不继承旧任务的具体诊断。"""
    if relation.get("relation") not in {"same_task", "same_project_new_task"}:
        return {"entities": extract_structure_entities(new_structure)}
    merged = old_project.copy() if isinstance(old_project, dict) else {}
    old_entities = merged.get("entities", {}) or {}
    new_entities = extract_structure_entities(new_structure)
    out = {"object": new_entities.get("object", "") or old_entities.get("object", "")}
    for key in ["events", "messages", "costumes", "variables", "targets", "block_terms"]:
        out[key] = set(old_entities.get(key, set())) | set(new_entities.get(key, set()))
    return {"entities": out}


# =========================
# 4.6 Qwen-VL：学生作品运行视频分析
# =========================

def save_uploaded_video_to_temp(video_file):
    """
    将 Streamlit 上传的视频临时保存到本地，供 OpenCV 抽取关键帧。
    """
    suffix = os.path.splitext(video_file.name or "uploaded_video.mp4")[1] or ".mp4"
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    video_file.seek(0)
    temp.write(video_file.getvalue())
    temp.flush()
    temp.close()
    video_file.seek(0)
    return temp.name


def pil_image_to_data_url(pil_img, image_format="JPEG"):
    """
    将 PIL 图片转为 Qwen-VL 可读取的 data URL。
    """
    buffer = io.BytesIO()
    pil_img = pil_img.convert("RGB")
    pil_img.thumbnail((960, 540))
    pil_img.save(buffer, format=image_format, quality=85)
    encoded = base64.b64encode(buffer.getvalue()).decode("utf-8")
    mime = "image/jpeg" if image_format.upper() == "JPEG" else "image/png"
    return f"data:{mime};base64,{encoded}"


def decide_video_frame_count(duration):
    """
    根据视频时长自动决定抽取几帧。
    """
    if duration <= 15:
        return 5
    elif duration <= 30:
        return 6
    elif duration <= 60:
        return 8
    else:
        return 10


def extract_video_keyframes(video_file):
    """
    自动根据视频时长均匀抽取关键帧。
    短视频也能抽到开始、中间、结尾，避免12秒视频只抽2-3帧。
    返回：[{time, image, data_url}, ...]
    """
    try:
        import cv2
    except Exception:
        st.warning("当前环境没有安装 opencv-python-headless，无法抽取视频关键帧。")
        return []

    temp_video_path = save_uploaded_video_to_temp(video_file)
    frames = []

    try:
        cap = cv2.VideoCapture(temp_video_path)
        if not cap.isOpened():
            return []

        fps = cap.get(cv2.CAP_PROP_FPS) or 25
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
        duration = total_frames / fps if fps else 0

        if duration <= 0:
            candidate_times = [0]
        else:
            frame_count = decide_video_frame_count(duration)

            if frame_count <= 1:
                candidate_times = [0]
            else:
                candidate_times = [
                    round(i * duration / (frame_count - 1), 2)
                    for i in range(frame_count)
                ]

        for second in candidate_times:
            cap.set(cv2.CAP_PROP_POS_MSEC, second * 1000)
            success, frame = cap.read()

            if not success or frame is None:
                continue

            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            pil_img = PILImage.fromarray(frame_rgb)

            frames.append({
                "time": second,
                "image": pil_img,
                "data_url": pil_image_to_data_url(pil_img)
            })

        cap.release()

    finally:
        try:
            os.remove(temp_video_path)
        except Exception:
            pass

    return frames


def analyze_video_frame_with_qwen_vl(frame_item, task_context=""):
    """
    使用 Qwen-VL 分析单张关键帧中的图形化编程运行效果。
    """
    if ocr_client is None:
        return "未配置 DASHSCOPE_API_KEY，无法调用 Qwen-VL 分析视频关键帧。"

    prompt = f"""
你是小学图形化编程作品运行效果观察助手。
请只根据这一帧画面，判断学生作品运行时呈现出的可见现象。

观察重点：
1. 是否能看到图形化编程舞台或作品运行画面；
2. 角色是否存在、位置是否正常；
3. 是否能看到分数、倒计时、提示文字、胜利/失败等信息；
4. 是否能看出角色移动、碰撞、场景切换、交互反馈等运行效果；
5. 是否存在明显异常，例如画面空白、角色消失、卡住、变量显示异常。

注意：
1. 不要凭空猜测代码。
2. 如果单帧无法判断动态过程，请明确写“单帧无法确认”。
3. 语言简洁，适合放入教师评价表。

课堂任务或评价量表背景：
{task_context[:1000]}

当前关键帧时间：第 {frame_item.get('time', 0)} 秒
"""
    try:
        response = ocr_client.chat.completions.create(
            model=QWEN_VL_MODEL,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": frame_item["data_url"]}}
                    ]
                }
            ],
            temperature=0.1,
            max_tokens=700
        )

        raw_text = response.choices[0].message.content or ""
        return clean_answer(raw_text)

    except Exception as e:
        return f"该关键帧调用 Qwen-VL 分析失败：{e}"


def summarize_video_analysis_with_qwen_vl(frame_observations, task_context=""):
    """
    汇总多个关键帧观察结果，形成作品运行视频分析。
    """
    if not frame_observations:
        return "未获得有效视频关键帧，无法进行运行画面分析。"

    if ocr_client is None:
        return "未配置 DASHSCOPE_API_KEY，无法调用 Qwen-VL 汇总视频分析。"

    observations_text = "\n".join([
        f"第{item['time']}秒：{item['observation']}" for item in frame_observations
    ])

    prompt = f"""
你是小学五年级图形化编程作品运行视频辅助评价助手。
请根据多个关键帧观察结果，汇总学生作品的运行表现。

要求：
1. 重点判断作品是否有可见运行效果；
2. 说明角色、变量、提示文字、场景变化、交互反馈等是否可见；
3. 区分“已经观察到”和“关键帧无法确认”；
4. 不替代教师正式评分，只作为教师复核参考；
5. 输出简洁，便于写入 Excel。

请严格按以下格式输出：
运行画面分析：
可确认的运行效果：
无法确认或需教师复核：

课堂任务或评价量表背景：
{task_context[:1200]}

关键帧观察结果：
{observations_text}
"""

    try:
        response = ocr_client.chat.completions.create(
            model=QWEN_VL_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
            max_tokens=1200
        )

        raw_text = response.choices[0].message.content or ""
        return clean_answer(raw_text)

    except Exception as e:
        return f"视频关键帧汇总分析失败：{e}"


def analyze_project_video_with_qwen_vl(video_file, task_context=""):
    """
    完整视频分析流程：上传视频 → 抽取关键帧 → Qwen-VL 分析单帧 → 汇总运行表现。
    """
    frames = extract_video_keyframes(video_file)

    if not frames:
        return {
            "video_file": video_file.name if video_file else "",
            "frame_count": 0,
            "frame_observations": [],
            "summary": "未能抽取有效关键帧，无法分析运行视频。"
        }

    frame_observations = []
    for frame_item in frames:
        observation = analyze_video_frame_with_qwen_vl(frame_item, task_context=task_context)
        frame_observations.append({
            "time": frame_item.get("time", 0),
            "observation": observation
        })

    summary = summarize_video_analysis_with_qwen_vl(
        frame_observations,
        task_context=task_context
    )

    return {
        "video_file": video_file.name if video_file else "",
        "frame_count": len(frames),
        "frame_observations": frame_observations,
        "summary": summary
    }


def normalize_file_stem(file_name):
    """
    统一文件名主干，用于把学生 .sb3 和运行视频进行粗略匹配。
    例如：张三.sb3、张三.mp4 可以自动匹配。
    """
    stem = os.path.splitext(os.path.basename(file_name or ""))[0]
    stem = re.sub(r"[\s_\-（）()]+", "", stem)
    return stem.lower()


def match_video_for_sb3(sb3_file_name, video_files):
    """
    根据文件名匹配对应运行视频。
    优先完全同名主干；如果只有一个视频，也默认匹配给当前作品。
    """
    if not video_files:
        return None

    if len(video_files) == 1:
        return video_files[0]

    sb3_stem = normalize_file_stem(sb3_file_name)
    for video_file in video_files:
        if normalize_file_stem(video_file.name) == sb3_stem:
            return video_file

    for video_file in video_files:
        video_stem = normalize_file_stem(video_file.name)
        if sb3_stem and (sb3_stem in video_stem or video_stem in sb3_stem):
            return video_file

    return None


# =========================
# 5. 调用 DeepSeek
# =========================

def build_api_messages(
    role: str,
    messages: list,
    knowledge_context: str = "",
    current_topic: str = "",
    theme_reference_context: str = "",
    answer_mode: str = "normal",
) -> list:
    role_prompt = build_role_prompt(role)
    system_content = SYSTEM_PROMPT + "\n\n" + role_prompt

    # 图形化编程通用知识库只用于学生端学习支持。
    if role == "学生端":
        system_content += (
            "\n\n"
            + build_student_knowledge_prompt(
                current_theme=current_topic,
                knowledge_context=knowledge_context
            )
            + "\n\n"
            + build_theme_mode_prompt(
                mode=answer_mode,
                theme_reference_context=theme_reference_context,
            )
        )

    api_messages = [
        {"role": "system", "content": system_content}
    ]

    for msg in messages:
        api_messages.append({
            "role": msg["role"],
            "content": msg["content"]
        })

    return api_messages


def call_deepseek_full(
    role: str,
    messages: list,
    knowledge_context: str = "",
    current_topic: str = "",
    theme_reference_context: str = "",
    answer_mode: str = "normal",
) -> str:
    """
    非流式完整生成。教师端长回答使用这个函数，避免界面先显示半截内容。
    """
    api_messages = build_api_messages(
        role,
        messages,
        knowledge_context=knowledge_context,
        current_topic=current_topic,
        theme_reference_context=theme_reference_context,
        answer_mode=answer_mode,
    )
    max_tokens = 12000 if role == "教师端" else 1600

    response = client.chat.completions.create(
        model="deepseek-chat",
        messages=api_messages,
        temperature=0.25,
        max_tokens=max_tokens,
        stream=False
    )

    raw_answer = response.choices[0].message.content or ""
    answer = postprocess_answer(role, raw_answer)

    finish_reason = getattr(response.choices[0], "finish_reason", "")
    if finish_reason == "length":
        answer += "\n\n提示：本次回答内容较长，可能仍有部分内容被模型长度限制截断。可以让智能体继续补充后续内容。"

    return answer


def stream_deepseek(
    role: str,
    messages: list,
    placeholder,
    scroll_holder=None,
    knowledge_context: str = "",
    current_topic: str = "",
    max_tokens_override: int = None,
    theme_reference_context: str = "",
    answer_mode: str = "normal",
) -> str:
    """
    学生端使用流式调用，让回答逐步出现。
    教师端建议使用 call_deepseek_full()，保证长教学设计完整显示。
    """
    api_messages = build_api_messages(
        role,
        messages,
        knowledge_context=knowledge_context,
        current_topic=current_topic,
        theme_reference_context=theme_reference_context,
        answer_mode=answer_mode,
    )
    max_tokens = max_tokens_override or (1600 if role == "学生端" else 12000)

    response = client.chat.completions.create(
        model="deepseek-chat",
        messages=api_messages,
        temperature=0.25,
        max_tokens=max_tokens,
        stream=True
    )

    full_answer = ""
    last_update_time = 0
    last_scroll_time = 0

    for chunk in response:
        try:
            delta = chunk.choices[0].delta.content or ""
        except Exception:
            delta = ""

        if not delta:
            continue

        full_answer += delta
        now = time.time()

        if now - last_update_time > 0.08:
            # 流式中间态只做基础清理，不做编号重排，避免半句话被误处理成单独标题
            temp_answer = clean_answer(full_answer)
            placeholder.markdown(
                build_chat_bubble_html("assistant", temp_answer),
                unsafe_allow_html=True
            )
            last_update_time = now

        if now - last_scroll_time > 0.65:
            try:
                scroll_to_bottom(scroll_holder, smooth=False)
            except Exception:
                pass
            last_scroll_time = now

    answer = postprocess_answer(role, full_answer)

    # 最后必须用完整 full_answer 再渲染一次，避免界面停留在流式输出的中间状态。
    placeholder.markdown(
        build_chat_bubble_html("assistant", answer),
        unsafe_allow_html=True
    )

    try:
        scroll_to_bottom(scroll_holder, smooth=False)
    except Exception:
        pass

    return answer


# =========================
# 5.5 开场白问题
# =========================

def get_starter_questions(role: str):
    if role == "教师端":
        return [
            ("教学设计", "请帮我设计一节《打地鼠》图形化编程教学设计。"),
            ("任务单", "请帮我生成一份适合五年级学生的图形化编程任务单。"),
            ("调试支持", "请帮我整理学生常见的图形化编程调试问题。")
        ]

    return [
        ("任务分析", "我想做一个打地鼠小游戏，可以帮我分析一下怎么做吗？"),
        ("调试帮助", "我的角色只动了一次，应该先检查哪里？"),
        ("变量问题", "我想让分数增加，但分数没有变化，可以怎么排查？"),
        ("图片识别", "我上传了一张截图，请告诉我哪有问题。")
    ]


def set_quick_prompt(prompt_text: str):
    st.session_state.quick_prompt = prompt_text


# =========================
# 6. 侧边栏：身份入口
# =========================

st.sidebar.title("身份选择")

role_display_map = {"学生端": "学生", "教师端": "老师"}
user_role = st.sidebar.radio(
    "请选择你的身份",
    ["学生端", "教师端"],
    format_func=lambda role: role_display_map.get(role, role)
)


TEACHER_PASSWORD = get_secret_or_env("TEACHER_PASSWORD", "teacher123")

if user_role == "教师端":
    password = st.sidebar.text_input("请输入教师端密码", type="password")
    if password != TEACHER_PASSWORD:
        st.warning("老师身份需要密码。学生请选择左侧的“学生”。")
        st.stop()

    st.sidebar.divider()
    st.sidebar.markdown("### 学生端知识库")
    if KNOWLEDGE_LOAD_ERROR:
        st.sidebar.warning(KNOWLEDGE_LOAD_ERROR)
    else:
        st.sidebar.success(f"已加载 {len(KNOWLEDGE_BASE_DF)} 条知识")
        st.sidebar.caption("文件：scratch_knowledge.xlsx")


# 身份切换时清空当前页面对话
if "current_role" not in st.session_state:
    st.session_state.current_role = user_role

if st.session_state.current_role != user_role:
    st.session_state.current_role = user_role
    st.session_state.messages = []
    st.session_state.quick_prompt = ""
    st.session_state.pasted_image = None
    st.session_state.current_debug_context = {}
    st.session_state.project_debug_context = {}
    st.session_state.previous_debug_contexts = []


# 初始化聊天记录
if "messages" not in st.session_state:
    st.session_state.messages = []


# 初始化学生会话标识
if "student_session_id" not in st.session_state:
    st.session_state.student_session_id = ""

# 初始化连续截图调试上下文
if "current_debug_context" not in st.session_state:
    st.session_state.current_debug_context = {}
if "project_debug_context" not in st.session_state:
    st.session_state.project_debug_context = {}
if "previous_debug_contexts" not in st.session_state:
    st.session_state.previous_debug_contexts = []


# 初始化快捷问题
if "quick_prompt" not in st.session_state:
    st.session_state.quick_prompt = ""

if "last_sent_paste_hash" not in st.session_state:
    st.session_state.last_sent_paste_hash = ""


# =========================
# 7. 页面主体
# =========================

st.markdown(
    """
    <div class="main-card">
        <div class="title-row">
            <div class="title-icon">🐱</div>
            <div class="title-text">
                <div class="app-title">图形化编程学习助手</div>
                <p class="app-desc">提问时说清楚：想实现什么、做到了哪一步、遇到了什么问题。</p>
            </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True
)

if user_role == "学生端":
    col1, col2 = st.columns([1, 1])

    with col1:
        group_choice = st.selectbox(
            "小组号",
            ["第1组", "第2组", "第3组", "第4组", "第5组", "第6组", "第7组", "第8组", "第9组",  "第10组", "第11组", "第12组","其他小组"],
            index=0
        )

    with col2:
        topic_choice = st.selectbox(
            "当前主题",
            ["海底世界", "猫捉老鼠",  "牛顿的苹果","打地鼠",  "其他主题"],
            index=0
        )

    if group_choice == "其他小组":
        group_no = st.text_input("请输入小组号", placeholder="如：第9组")
    else:
        group_no = group_choice

    if topic_choice == "其他主题":
        topic = st.text_input("请输入当前主题", placeholder="如：海底世界")
    else:
        topic = topic_choice

    # 学生端不再要求填写姓名，减少小学生打字负担
    student_name = ""

    # 只根据“小组号 + 当前主题”判断是否开启新对话
    # 小组号或当前主题任意一个发生变化，就清空当前页面对话
    # 注意：这不会删除已经保存到 logs/chat_logs.csv 的历史记录
    current_student_session = f"{group_no.strip()}_{topic.strip()}"

    if group_no.strip() and topic.strip():
        if current_student_session != st.session_state.student_session_id:
            st.session_state.student_session_id = current_student_session
            st.session_state.messages = []
            st.session_state.quick_prompt = ""
            st.session_state.pasted_image = None
            st.session_state.current_debug_context = {}
            st.session_state.project_debug_context = {}
            st.session_state.previous_debug_contexts = []

else:
    student_name = ""
    group_no = ""
    topic = ""


# =========================
# 7.5 开场白问题区域
# =========================

if not st.session_state.messages:
    render_chat_bubble(
        "assistant",
        "你好，我可以帮你分析图形化编程任务、检查程序问题，也可以识别你上传的程序截图。你可以直接提问，也可以从下面选一个问题开始。"
    )

    st.markdown('<div class="starter-area">', unsafe_allow_html=True)
    starter_questions = get_starter_questions(user_role)

    for i, item in enumerate(starter_questions):
        label, question = item
        st.button(
            question,
            key=f"starter_{user_role}_{i}",
            on_click=set_quick_prompt,
            args=(question,)
        )
    st.markdown('</div>', unsafe_allow_html=True)


# =========================
# 8. 显示历史聊天消息
# =========================

for msg in st.session_state.messages:
    render_chat_bubble(
        msg["role"],
        msg["content"],
        msg.get("image_base64", "")
    )

# 粘贴截图初始化：只负责暂存，不自动发送
if "pasted_image" not in st.session_state:
    st.session_state.pasted_image = None

send_pasted_only = False

# =========================
# 9. 聊天输入框：支持文字 + 图片附件 + 粘贴截图
# =========================

if user_role == "学生端":
    input_placeholder = "请输入问题，或点击附件上传图片。"
else:
    input_placeholder = "请输入教学需求，或点击附件上传图片。"

# 粘贴截图：只暂存，不自动发送
# 学生可以先粘贴图片，再在聊天框中输入问题发送；
# 也可以不输入问题，直接点击“发送截图”。

prompt = st.chat_input(
    input_placeholder,
    accept_file=True,
    file_type=["png", "jpg", "jpeg"]
)

# 固定在底部、靠近聊天输入框的粘贴截图工具条
# 保留原生 st.chat_input 的样式；截图按钮放在输入框下方同一底部区域。
if user_role in ["学生端", "教师端"]:
    with st.container(key="paste_toolbar"):
        # 这里始终定义 preview_col 和 clear_col，避免刚粘贴图片的这一轮中变量未定义。
        paste_col, send_img_col, preview_col, clear_col, tip_col = st.columns([0.8, 0.75, 0.28, 0.22, 4.9])

        with paste_col:
            pasted_result = paste_image_button(
                label="📋 粘贴截图",
                key="paste_image_button"
            )

        if pasted_result.image_data is not None:
            # paste_image_button 有时会在 rerun 后继续返回上一张图。
            # 用哈希避免“发送后清空了，又被组件自动塞回来”。
            try:
                _buf = BytesIO()
                pasted_result.image_data.save(_buf, format="PNG")
                _paste_hash = base64.b64encode(_buf.getvalue()).decode("utf-8")[:80]
            except Exception:
                _paste_hash = str(time.time())

            if _paste_hash != st.session_state.get("last_sent_paste_hash", ""):
                st.session_state.pasted_image = pasted_result.image_data
                st.session_state.current_paste_hash = _paste_hash

        with send_img_col:
            send_pasted_only = st.button(
                "发送截图",
                key="send_pasted_only",
                disabled=st.session_state.pasted_image is None
            )

        if st.session_state.pasted_image is not None:
            with preview_col:
                st.image(
                    st.session_state.pasted_image,
                    width=46
                )

            with clear_col:
                if st.button("×", key="clear_pasted_image"):
                    st.session_state.pasted_image = None
                    st.rerun()

            with tip_col:
                st.markdown(
                    '<div class="pending-image-box">已粘贴，可输入问题后发送，或直接发送截图。</div>',
                    unsafe_allow_html=True
                )
        else:
            with tip_col:
                st.empty()

if prompt:
    # Streamlit 版本不同，st.chat_input 可能返回字符串，也可能返回包含 text/files 的对象
    if isinstance(prompt, str):
        user_input = prompt
        uploaded_image = None
    else:
        user_input = getattr(prompt, "text", "") or ""
        uploaded_image = prompt.files[0] if getattr(prompt, "files", None) else None

elif send_pasted_only:
    user_input = ""
    uploaded_image = None

elif st.session_state.quick_prompt:
    user_input = st.session_state.quick_prompt
    uploaded_image = None
    st.session_state.quick_prompt = ""

else:
    user_input = None
    uploaded_image = None

# 本轮要发送的图片来源：
# ① 如果聊天框附件上传了图片，优先使用附件图片；
# ② 如果没有附件图片，但学生之前粘贴了截图，则随本轮文字一起发送粘贴截图；
# ③ 只粘贴截图但没有点击聊天框发送键时，不会触发智能体。
pasted_image_to_send = None
image_for_this_turn = uploaded_image

if user_input is not None and uploaded_image is None:
    if st.session_state.pasted_image is not None:
        pasted_image_to_send = st.session_state.pasted_image

if image_for_this_turn is None and pasted_image_to_send is not None:
    buffer = BytesIO()
    pasted_image_to_send.save(buffer, format="PNG")
    buffer.seek(0)
    buffer.name = "pasted_image.png"
    buffer.type = "image/png"
    image_for_this_turn = buffer

# 只有真正输入文字、点击开场问题、聊天框发送文字并携带待发送截图，或点击“发送截图”时，才调用智能体
if user_input is not None:
    if user_role == "学生端":
        if not group_no.strip() or not topic.strip():
            st.warning("请先选择小组号和当前主题，再进行提问。")
            st.stop()

    has_text_question = bool(str(user_input or "").strip())
    is_image_only_turn = image_for_this_turn is not None and not has_text_question

    if has_text_question:
        display_user_input = str(user_input).strip()
        api_user_input = str(user_input).strip()
    elif is_image_only_turn and user_role == "学生端":
        display_user_input = "请帮我分析这张程序截图中存在哪些问题。"
        api_user_input = (
            "学生本轮只上传了程序截图，没有输入文字问题。"
            "请依据截图中能够确认的程序结构，分析其中存在的问题。"
        )
    elif is_image_only_turn:
        display_user_input = "请分析这张截图。"
        api_user_input = "请结合截图内容进行分析。"
    else:
        display_user_input = str(user_input or "").strip()
        api_user_input = display_user_input
    uploaded_image_name = ""
    uploaded_image_path = ""
    uploaded_image_base64 = ""
    ocr_text = ""
    screenshot_analysis = ""
    screenshot_hard_constraints = ""
    image_only_direct_answer = ""
    screenshot_relation = {}
    active_debug_context_prompt = ""
    current_question_goal = infer_question_goal(api_user_input)
    answer_mode = classify_student_answer_mode(
        (str(user_input).strip() if has_text_question else api_user_input),
        has_screenshot=(image_for_this_turn is not None),
    )
    theme_reference_context = build_theme_reference_context(
        topic,
        current_object="",
        student_question=(str(user_input).strip() if has_text_question else api_user_input),
    )
    structured_screenshot_diagnosis = {}

    # 如果上传了图片，先保存原图，再进行 OCR 文字识别
    if image_for_this_turn is not None:
        uploaded_image_name, uploaded_image_path = save_uploaded_image_file(image_for_this_turn)
        image_for_this_turn.seek(0)
        uploaded_image_base64 = base64.b64encode(image_for_this_turn.getvalue()).decode("utf-8")
        image_for_this_turn.seek(0)

        with st.spinner("正在识别图片中的文字，请稍等……"):
            ocr_text = recognize_image_text(image_for_this_turn)

        # OCR只能识别文字，不能可靠判断积木嵌套和多段脚本关系。
        # 学生端额外调用Qwen-VL提取截图结构证据，再交给DeepSeek组织教学提示。
        if user_role == "学生端":
            image_for_this_turn.seek(0)
            with st.spinner("正在分析截图中的程序结构，请稍等……"):
                screenshot_analysis = analyze_student_screenshot_with_qwen_vl(
                    image_for_this_turn,
                    student_question=api_user_input,
                )
            image_for_this_turn.seek(0)

            # 先读取所有可见脚本，再结合教师基础版目标进行第二阶段诊断。
            # 不再使用容易误判的主题硬编码答案。
            current_structure = get_screenshot_structure(screenshot_analysis)
            theme_reference_context = build_theme_reference_context(
                topic,
                current_object=current_structure.get("current_object", ""),
                student_question=(str(user_input).strip() if has_text_question else api_user_input),
            )
            structured_screenshot_diagnosis = diagnose_screenshot_with_teacher_reference(
                student_question=(str(user_input).strip() if has_text_question else ""),
                screenshot_structure=current_structure,
                ocr_text=ocr_text,
                theme_reference_context=theme_reference_context,
                answer_mode=answer_mode,
            )
            image_only_direct_answer = render_grounded_student_answer(structured_screenshot_diagnosis)
            screenshot_hard_constraints = json.dumps(
                {
                    "source": "teacher_reference_grounded_diagnosis",
                    "mode": answer_mode,
                    "diagnosis": structured_screenshot_diagnosis,
                    "rules": [
                        "教师基础版只用于理解基础功能，不要求学生代码逐块相同",
                        "普通调试只选择一个最相关问题并给一个具体改法",
                        "证据不足时必须追问，不得猜测脚本冲突或删除脚本",
                        "只有学生明确询问创新时才提供拓展建议",
                    ],
                },
                ensure_ascii=False,
            )

            # 判断本轮新截图与最近一次截图调试上下文的关系。
            previous_context = st.session_state.get("current_debug_context", {}) or {}
            screenshot_relation = compare_debug_contexts(
                previous=previous_context,
                new_structure=current_structure,
                new_question=(str(user_input).strip() if has_text_question else api_user_input),
            )

            # 同一作品的新任务不能继承旧诊断；同一任务可以保留作品背景，
            # 但本轮截图和本轮问题仍然具有最高优先级。
            if previous_context and not screenshot_relation.get("inherit_debug_diagnosis", False):
                history = st.session_state.get("previous_debug_contexts", [])
                history.append(previous_context)
                st.session_state.previous_debug_contexts = history[-8:]

            st.session_state.project_debug_context = update_project_context(
                st.session_state.get("project_debug_context", {}) or {},
                current_structure,
                screenshot_relation,
            )
            st.session_state.current_debug_context = {
                "structure": current_structure,
                "question": (str(user_input).strip() if has_text_question else api_user_input),
                "goal": current_question_goal,
                "relation": screenshot_relation,
                "ocr_text": ocr_text,
                "hard_constraints": screenshot_hard_constraints,
                "teacher_theme_reference": theme_reference_context,
                "structured_diagnosis": structured_screenshot_diagnosis,
                "answer_mode": answer_mode,
            }
            active_debug_context_prompt = build_active_debug_context_prompt(
                st.session_state.current_debug_context
            )

        # 不把“已上传图片：xxx”重复显示在聊天文字里；
        # 图片会直接嵌入用户消息气泡中。
        api_user_input = f"""
【本轮用户最新问题】
{api_user_input}

【本轮截图结构分析（视觉证据）】
{screenshot_analysis or "未获得截图结构分析。"}

【本轮截图文字识别结果】
{ocr_text}

【高置信度诊断约束】
{screenshot_hard_constraints or "本轮无额外硬性约束。"}

【学生本轮目标提取】
触发方式：{current_question_goal.get("trigger") or "未确认"}
期望动作：{"、".join(current_question_goal.get("actions", [])) or "未确认"}

【与最近一次截图的关系】
{screenshot_relation or {"relation": "new_context", "reason": "没有上一轮截图上下文"}}

【教师基础版主题参照】
{theme_reference_context}

【当前回答模式】
{answer_mode}

请根据“本轮用户最新问题”、截图证据和教师基础版目标作答。

注意：
1. 只围绕本轮最新输入回答，不要重复回答历史问题。
1.1 你正在直接回复当前学生。成文时一律使用“你”“你的程序”“你的这段脚本”，不得用“学生”“该学生”“学生把……”等第三人称指代当前对话者。例如应写“你把点击触发的敲击动作放在了……”，不要写“学生把点击触发的敲击动作放在了……”。
2. 如果本轮有文字问题，必须以文字问题为主，结合截图中与该问题直接相关的证据回答，不要另行扩展其他问题。
3. 如果本轮只有截图、没有文字问题，默认帮助学生分析截图中的程序问题：先指出截图能够直接确认的最主要问题，并给出一个修改方向；如果截图还明确显示第二个与程序运行直接相关的问题，可以再简要指出，但不要一次罗列大量可能原因。
4. 截图结构分析的优先级高于OCR文字和知识库；不得用常见错误替代截图中已经显示的事实。
5. 在回答“缺少等待、缺少循环、缺少条件”等结论前，必须核对截图结构分析是否明确显示该积木已经存在。
6. 必须先读取截图中的全部可见独立脚本，逐段分析每段脚本的事件、循环、条件和动作，再综合检查它们是否同时修改造型、显示状态、位置、方向、大小或变量。不得只分析前两段脚本。
7. 不要向学生逐字展示OCR结果或后台截图结构分析，除非学生明确要求识别文字。
8. 如果结构分析与OCR矛盾、截图连接关系不清、关键积木被遮挡，或无法判断学生想实现的效果，不要猜测；说明当前截图能确认的内容，并只追问一个关键问题。
9. 截图中能够明确判断错误时，必须在简短分析后给出“修改后的关键程序”框。程序框只包含与当前问题直接相关的一段脚本，不包含整个作品。
10. 修改方向必须告诉学生怎样修改现有程序，例如“把循环执行改成一次执行”“把等待中的条件改成固定秒数”或“把条件放入如果……那么”。不得只说“删掉脚本”“只保留另一段脚本”“重新写”。
11. 必须优先采用“高置信度诊断约束”，不得用一般经验替换其中已确认的问题，也不得输出其中明确禁止的误判。截图没有显示“停止”积木、状态变量或暂停机制时，不得自行建议添加这些结构。
12. 只有截图中没有任何可确认结构错误时，才询问学生想实现什么效果；不能在已有明确错误时先追问目标。
13. 只有截图证据足以确定正确连接关系时，才使用 [[PROGRAM]] 和 [[/PROGRAM]] 展示局部程序，并按积木嵌套关系缩进。程序框中的已有积木文字必须来自截图结构分析的原始 text 字段，不得根据常识重新拼写；证据不足时不得编造程序，应只追问一个关键问题。程序框前说明原因和具体修改，程序框后鼓励学生运行验证。
14. “为什么一直闪现”和“打到后没有哭”属于不同问题，必须分别分析；不得因为截图结构相似就复用同一回答。
15. 截图证据充分时，回答结构为：简短说明原因；说明怎样修改；用一个局部程序框呈现正确连接顺序；鼓励学生运行验证。截图证据不足时，不得强行输出程序框。
16. 必须先从学生原话中识别“由什么触发”和“想发生什么动作”。学生说“点击锤子时左转再右转”，就不能擅自改成“碰到地鼠时触发”。
17. 同一作品中的新角色或新功能属于新调试任务，只能继承作品背景，不能继承上一任务的错误结论。
18. 若学生明确补充或纠正目标，以最新原话为准；不得继续沿用模型上一轮的误解。
19. 学生目标与截图结构不矛盾时，应直接说明需要移动、保留或新增哪些现有积木；不要引入截图和问题中没有出现的碰撞、广播、变量或停止脚本。
20. 当回答“x坐标/y坐标/方向/等待时间/重复次数/变量值应该填多少”这类问题时，必须同时说明数值依据，不能只给答案。优先用截图中已有数值做简短计算，例如“现在 x 是 -98，前面把 x 增加 130，所以停下来的位置是 32”。
21. 如果学生的目标仍有两种以上合理理解，只追问一个确认问题，不要先输出完整程序。
"""

    # 没有上传新截图时，如果学生是在补充、纠正或追问最近一次截图，
    # 继续携带最近一次有效截图结构；避免只看一句“是我控制”而脱离上下文猜测。
    if image_for_this_turn is None and user_role == "学生端":
        active_context = st.session_state.get("current_debug_context", {}) or {}
        related_to_recent_screenshot = is_text_related_to_context(api_user_input, active_context)
        if active_context and related_to_recent_screenshot:
            active_debug_context_prompt = build_active_debug_context_prompt(active_context)
            followup_flag = is_followup_or_correction(api_user_input)
            api_user_input = f"""
【学生本轮最新输入】
{api_user_input}

{active_debug_context_prompt}

【本轮是否像补充或纠正】
{followup_flag}

请先结合最近截图理解学生最新输入。学生最新输入优先级最高：
1. 若是在补充或纠正目标，先更新对目标的理解，再回答，不得重复旧结论。
2. 若学生说“不是……”“是我控制”“我的意思是……”，必须把它理解为对最近截图任务的修正。
3. 若本轮已经足以确定触发方式和动作，直接给出对应的局部修改；若仍有歧义，只问一个确认问题。
4. 不得把知识库中的常见场景替代最近截图中的实际角色、事件和积木。
""".strip()

    # 页面上显示用户文字和对应图片，历史消息中也保留这张图片
    st.session_state.messages.append({
        "role": "user",
        "content": display_user_input,
        "api_content": api_user_input,
        "image_base64": uploaded_image_base64
    })

    render_chat_bubble("user", display_user_input, uploaded_image_base64)

    # 用户消息出现后，尽量自动滚动到底部
    scroll_holder = st.empty()
    try:
        scroll_to_bottom(scroll_holder, smooth=True)
    except Exception:
        pass

    with st.spinner("正在生成中，请稍等……"):
        try:
            # 历史记录只放“用户真实显示文字”和助手回答，避免上一轮 OCR 长文本反复干扰本轮问题。
            # 本轮如果有截图，只把本轮 OCR 拼进最后一条 user 消息。
            # 截图调试必须只依据本轮问题和本轮截图，避免上一轮错误回复继续影响本轮。
            if image_for_this_turn is not None:
                messages_for_api = [
                    {
                        "role": "user",
                        "content": api_user_input
                    }
                ]
            else:
                messages_for_api = [
                    {
                        "role": msg["role"],
                        "content": msg["content"]
                    }
                    for msg in st.session_state.messages[:-1]
                ] + [
                    {
                        "role": "user",
                        "content": api_user_input
                    }
                ]

            # 学生端根据“本轮问题 + OCR结果”检索最相关的1—3条知识。
            knowledge_context = ""
            if user_role == "学生端" and image_for_this_turn is None:
                # 对最近截图的补充/纠正，优先使用截图上下文，不调用知识库，避免通用案例覆盖学生当前目标。
                active_context = st.session_state.get("current_debug_context", {}) or {}
                followup_to_screenshot = is_text_related_to_context(
                    (str(user_input).strip() if has_text_question else api_user_input),
                    active_context,
                )
                if not followup_to_screenshot:
                    matched_knowledge = retrieve_knowledge(
                        knowledge_df=KNOWLEDGE_BASE_DF,
                        student_question=(str(user_input).strip() if has_text_question else ""),
                        current_theme=topic,
                        ocr_text="",
                        top_k=3,
                    )
                    knowledge_context = build_knowledge_context(matched_knowledge)

            assistant_placeholder = st.empty()

            # 学生端保留流式输出；教师端使用完整生成，避免长教学设计在界面上显示半截。
            if user_role == "教师端":
                answer = call_deepseek_full(
                    user_role,
                    messages_for_api,
                    current_topic=topic,
                    theme_reference_context=theme_reference_context,
                    answer_mode=answer_mode,
                )
                assistant_placeholder.markdown(
                    build_chat_bubble_html("assistant", answer),
                    unsafe_allow_html=True
                )
            else:
                if image_for_this_turn is not None and image_only_direct_answer:
                    # 对少量可由截图直接确认的高置信度结构错误，使用确定性短答，
                    # 避免回答模型把正常的随机出现逻辑误判为错误。
                    answer = postprocess_answer("学生端", image_only_direct_answer)
                    assistant_placeholder.markdown(
                        build_chat_bubble_html("assistant", answer),
                        unsafe_allow_html=True
                    )
                else:
                    answer = stream_deepseek(
                        user_role,
                        messages_for_api,
                        assistant_placeholder,
                        scroll_holder,
                        knowledge_context=knowledge_context,
                        current_topic=topic,
                        max_tokens_override=(520 if image_for_this_turn is not None else None),
                        theme_reference_context=theme_reference_context,
                        answer_mode=answer_mode,
                    )

                # 如果截图调试回答仍把“删除整段脚本”当作首选方法，
                # 且本轮已有高置信度的具体修改建议，则替换为具体修改建议。
                if (
                    user_role == "学生端"
                    and image_for_this_turn is not None
                    and image_only_direct_answer
                    and any(
                        phrase in str(answer)
                        for phrase in [
                            "删掉脚本", "删除脚本", "删掉这段脚本", "删除这段脚本",
                            "只保留脚本", "只保留另一段", "重新写一遍", "重写程序"
                        ]
                    )
                ):
                    answer = postprocess_answer("学生端", image_only_direct_answer)
                    assistant_placeholder.markdown(
                        build_chat_bubble_html("assistant", answer),
                        unsafe_allow_html=True
                    )

            try:
                scroll_to_bottom(scroll_holder, smooth=True)
            except Exception:
                pass

            # 纯文字追问若仍与最近截图相关，更新“当前学生目标”，但保留最近截图结构。
            if user_role == "学生端" and image_for_this_turn is None:
                active_context = st.session_state.get("current_debug_context", {}) or {}
                raw_latest_question = str(user_input).strip() if has_text_question else api_user_input
                if active_context and is_text_related_to_context(raw_latest_question, active_context):
                    latest_goal = infer_question_goal(raw_latest_question)
                    merged_goal = dict(active_context.get("goal", {}) or {})
                    if latest_goal.get("trigger"):
                        merged_goal["trigger"] = latest_goal["trigger"]
                    if latest_goal.get("actions"):
                        merged_goal["actions"] = latest_goal["actions"]
                    merged_goal["raw"] = raw_latest_question
                    active_context["goal"] = merged_goal
                    active_context["question"] = raw_latest_question
                    st.session_state.current_debug_context = active_context

            st.session_state.messages.append({
                "role": "assistant",
                "content": answer
            })

            log_user_input = (
                str(user_input).strip()
                if has_text_question
                else display_user_input
            )

            if uploaded_image_name:
                log_user_input = f"{log_user_input}（上传图片：{uploaded_image_name}）"

            save_log(
                role=user_role,
                user_input=log_user_input,
                answer=answer,
                student_name=student_name,
                group_no=group_no,
                topic=topic,
                uploaded_image_name=uploaded_image_name,
                uploaded_image_path=uploaded_image_path,
                ocr_text=ocr_text,
                screenshot_analysis=screenshot_analysis,
                uploaded_image_base64=uploaded_image_base64
            )

            # 如果本轮使用了粘贴截图，发送后清空待发送区。
            # 发送后的图片已经进入聊天记录，不会消失。
            if pasted_image_to_send is not None:
                st.session_state.last_sent_paste_hash = st.session_state.get("current_paste_hash", "")
                st.session_state.pasted_image = None
                st.rerun()

        except Exception as e:
            st.error(f"调用失败：{e}")
# =========================
# 9.5 教师端：批量作品辅助评价
# =========================

if user_role == "教师端":
    st.markdown("### 批量作品辅助评价")

    with st.expander("上传学生作品和评价量表", expanded=False):
        rubric_file = st.file_uploader(
            "1.上传作品评价量表（txt / csv / xlsx）",
            type=["txt", "csv", "xlsx"],
            key="rubric_file"
        )
        reference_sb3_file = st.file_uploader(
            "2.上传教师基础版作品 .sb3（用于提供核心功能目标参照，可选但推荐）",
            type=["sb3"],
            key="reference_sb3_file"
        )

        reference_video_file = st.file_uploader(
            "3.上传教师基础版作品运行视频（用于提供运行效果参照，可选但推荐）",
            type=["mp4", "mov", "avi", "mkv", "webm"],
            key="reference_video_file"
        )

        sb3_files = st.file_uploader(
            "4.批量上传学生 .sb3 作品文件",
            type=["sb3"],
            accept_multiple_files=True,
            key="sb3_files"
        )

        video_files = st.file_uploader(
            "5.可选：批量上传学生作品运行视频",
            type=["mp4", "mov", "avi", "mkv", "webm"],
            accept_multiple_files=True,
            key="video_files"
        )


        st.caption("教师基础版视频用于提供标准运行效果参照；学生视频建议与 .sb3 文件同名，例如：张三.sb3 和 张三.mp4。若只上传一个学生视频，会默认匹配到当前作品。")

        start_batch_eval = st.button("开始批量辅助评价")

        if start_batch_eval:
            if rubric_file is None:
                st.warning("请先上传作品评价量表。")
                st.stop()

            if not sb3_files:
                st.warning("请至少上传一个 .sb3 作品文件。")
                st.stop()

            # 读取评价量表
            rubric_text = ""

            if rubric_file.name.endswith(".txt"):
                rubric_text = rubric_file.getvalue().decode("utf-8", errors="ignore")

            elif rubric_file.name.endswith(".csv"):
                df_rubric = pd.read_csv(rubric_file)
                rubric_text = df_rubric.to_string(index=False)

            elif rubric_file.name.endswith(".xlsx"):
                df_rubric = pd.read_excel(rubric_file)
                rubric_text = df_rubric.to_string(index=False)

            # 读取教师基础版作品。它只作为“核心功能目标参照”，不是标准答案代码。
            reference_analysis = None
            if reference_sb3_file is not None:
                reference_analysis = analyze_sb3_file(reference_sb3_file)

            # 分析教师基础版运行视频。它用于提供当前任务的可见运行效果参照。
            # 这里用 st.info / st.success 保留提示，避免 spinner 完成后提示消失，造成“没有分析教师视频”的误解。
            reference_video_analysis = None
            if reference_video_file is not None:
                st.info(f"正在分析教师基础版运行视频：{reference_video_file.name}")
                with st.spinner(f"正在分析教师基础版运行视频：{reference_video_file.name}"):
                    reference_video_analysis = analyze_project_video_with_qwen_vl(
                        reference_video_file,
                        task_context=(rubric_text + "\n这是教师基础版作品运行视频，用于说明本节课作品应呈现的基础运行效果和核心交互。"),
                    )
                st.success(f"教师基础版运行视频分析完成：{reference_video_file.name}")
            else:
                st.warning("未上传教师基础版运行视频，本次评价将主要参考教师基础版 .sb3、学生作品、学生视频和评价量表。")

            results = []

            with st.spinner("正在批量解析学生作品并生成辅助评价，请稍等……"):
                for sb3_file in sb3_files:
                    analysis = analyze_sb3_file(sb3_file)

                    matched_video = match_video_for_sb3(sb3_file.name, video_files)
                    video_analysis = None
                    if matched_video is not None:
                        st.info(f"正在分析学生作品运行视频：{matched_video.name}")
                        video_analysis = analyze_project_video_with_qwen_vl(
                            matched_video,
                            task_context=(
                                rubric_text
                                + "\n这是学生作品运行视频，请观察学生作品实际运行效果。"
                                + ("\n教师基础版运行视频分析参照：" + json.dumps(reference_video_analysis, ensure_ascii=False) if reference_video_analysis else "")
                            ),
                        )

                    eval_text = evaluate_project_with_rubric(
                        sb3_analysis=analysis,
                        rubric_text=rubric_text,
                        reference_analysis=reference_analysis,
                        video_analysis=video_analysis,
                        reference_video_analysis=reference_video_analysis
                    )

                    dimension_scores = parse_dimension_evaluations(eval_text)
                    running_analysis = extract_eval_field(eval_text, "运行画面分析")
                    if not running_analysis and video_analysis:
                        running_analysis = video_analysis.get("summary", "")
                    if not running_analysis:
                        running_analysis = "未上传运行视频，本次评价主要依据作品结构分析、教师基础作品和评价量表。"

                    row_result = {
                        "作品文件": sb3_file.name,
                        "教师基础视频": reference_video_file.name if reference_video_file is not None else "",
                        "匹配视频": matched_video.name if matched_video is not None else "",
                        "视频关键帧数": video_analysis.get("frame_count", "") if video_analysis else "",
                        "运行画面分析": running_analysis,
                        "核心功能完成情况": extract_eval_field(eval_text, "核心功能完成情况"),
                        "完成等级": normalize_completion_level(extract_eval_field(eval_text, "完成等级")),
                    }

                    # 根据教师上传的评价量表动态生成“各维度评分及依据”列
                    row_result.update(dimension_scores)

                    row_result.update({
                        "建议总分": extract_score_number(extract_eval_field(eval_text, "建议总分")),
                        "总评依据": extract_eval_field(eval_text, "总评依据") or extract_eval_field(eval_text, "评分依据"),
                        "改进建议": extract_eval_field(eval_text, "改进建议"),
                        "教师复核点": extract_eval_field(eval_text, "教师复核点"),
                        "视频逐帧观察": json.dumps(video_analysis.get("frame_observations", []), ensure_ascii=False) if video_analysis else "",

                        "角色数量": analysis.get("sprite_count", ""),
                        "角色名称": "、".join(analysis.get("sprites", [])),
                        "积木总数": analysis.get("block_count", ""),
                        "事件积木": analysis.get("event_blocks", ""),
                        "循环积木": analysis.get("loop_blocks", ""),
                        "条件积木": analysis.get("condition_blocks", ""),
                        "变量积木": analysis.get("variable_blocks", ""),
                        "广播积木": analysis.get("broadcast_blocks", ""),
                        "结构分析": analysis.get("summary", "")
                    })

                    results.append(row_result)
            result_df = pd.DataFrame(results)

            # 动态调整列顺序：基础信息 → 运行与核心功能 → 动态维度评分及依据 → 总评与复核 → 结构分析
            fixed_front_cols = [
                "作品文件", "教师基础视频", "匹配视频", "视频关键帧数", "运行画面分析",
                "核心功能完成情况", "完成等级"
            ]
            fixed_back_cols = [
                "建议总分", "总评依据", "改进建议", "教师复核点", "视频逐帧观察",
                "角色数量", "角色名称", "积木总数", "事件积木", "循环积木",
                "条件积木", "变量积木", "广播积木", "结构分析"
            ]
            dynamic_dimension_cols = [
                col for col in result_df.columns
                if col.endswith("评分及依据")
                and col not in fixed_front_cols
                and col not in fixed_back_cols
            ]
            ordered_cols = (
                [col for col in fixed_front_cols if col in result_df.columns]
                + dynamic_dimension_cols
                + [col for col in fixed_back_cols if col in result_df.columns]
            )
            other_cols = [col for col in result_df.columns if col not in ordered_cols]
            result_df = result_df[ordered_cols + other_cols]

            st.success("批量辅助评价已完成。")
            st.dataframe(result_df, use_container_width=True)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                result_df.to_excel(writer, index=False, sheet_name="辅助评价结果")

            st.download_button(
                label="下载批量评价结果 Excel",
                data=output.getvalue(),
                file_name="批量作品辅助评价结果.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
# =========================
# 10. 侧边栏：教师端操作区
# =========================

# 教师端只保留一个下载按钮：下载全部记录 Excel（含上传图片）
if user_role == "教师端":
    logs = load_logs()

    if not logs.empty:
        st.sidebar.divider()
        st.sidebar.markdown("### 对话记录")

        excel_data = create_excel_with_images(logs)
        st.sidebar.download_button(
            label="下载全部对话记录",
            data=excel_data,
            file_name="chat_logs_with_images.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.sidebar.divider()
        st.sidebar.markdown("### 对话记录")
        st.sidebar.info("暂无可下载的对话记录。")
